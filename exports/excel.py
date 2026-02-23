"""
Excel Export Module

Handles XLSX file generation for ID card data.
This module is READ-ONLY - it never mutates data.
"""
import logging
from io import BytesIO
from typing import Dict, Any, Optional
from dataclasses import dataclass

from django.http import HttpResponse
from django.db.models import QuerySet

from .utils import (
    get_text_fields,
    generate_export_filename,
    format_field_value,
    sort_cards_for_export,
    stream_file_response,
)

logger = logging.getLogger(__name__)


@dataclass
class ExcelExportResult:
    """Result of an Excel export operation."""
    success: bool
    message: str = ''
    response: Optional[HttpResponse] = None
    filename: str = ''
    row_count: int = 0


class ExcelExporter:
    """
    Handles Excel (XLSX) export operations.
    
    Features:
    - Exports only text fields (no images)
    - Auto-sizes columns
    - Applies consistent formatting
    - Freezes header row
    
    Usage:
        exporter = ExcelExporter()
        result = exporter.export_cards(table, cards)
        if result.success:
            return result.response
    """
    
    # Excel sheet name max length
    MAX_SHEET_NAME_LENGTH = 31
    
    # Column width limits
    MIN_COLUMN_WIDTH = 8
    MAX_COLUMN_WIDTH = 50
    
    # Maximum cards for Excel export (prevents OOM with large datasets)
    MAX_EXCEL_CARDS = 5000
    
    def export_cards(
        self,
        table,
        cards: QuerySet,
        uppercase_values: bool = True,
        status: str = ''
    ) -> ExcelExportResult:
        """
        Export cards to Excel format.
        
        Args:
            table: IDCardTable instance
            cards: QuerySet of IDCard instances
            uppercase_values: Convert values to uppercase
            
        Returns:
            ExcelExportResult with HttpResponse if successful
        """
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return ExcelExportResult(
                success=False,
                message='openpyxl library not installed. Run: pip install openpyxl'
            )
        
        if not cards.exists():
            return ExcelExportResult(
                success=False,
                message='No cards to export!'
            )
        
        try:
            # Get text fields only (exclude images)
            text_fields = get_text_fields(table.fields or [])
            
            if not text_fields:
                return ExcelExportResult(
                    success=False,
                    message='No text fields found in table configuration!'
                )
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = table.name[:self.MAX_SHEET_NAME_LENGTH]
            
            # Define styles
            styles = self._get_styles(Font, Alignment, Border, Side)
            
            # Track column widths
            column_widths = {}
            
            # Write header row
            headers = [f['name'] for f in text_fields]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = styles['header_font']
                cell.alignment = styles['header_alignment']
                cell.border = styles['border']
                column_widths[col_idx] = len(str(header)) + 2
            
            # Sort cards for export (Class → Name, or Name only)
            sorted_cards = sort_cards_for_export(list(cards[:self.MAX_EXCEL_CARDS]), table.fields)

            # Write data rows
            row_count = 0
            for row_idx, card in enumerate(sorted_cards, 2):
                field_data = card.field_data or {}
                
                for col_idx, field in enumerate(text_fields, 1):
                    value = field_data.get(field['name'], '')
                    formatted_value = format_field_value(value, uppercase=uppercase_values)
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=formatted_value)
                    cell.font = styles['data_font']
                    cell.alignment = styles['data_alignment']
                    cell.border = styles['border']
                    
                    # Track max width
                    current_width = min(len(formatted_value) + 2, self.MAX_COLUMN_WIDTH)
                    column_widths[col_idx] = max(
                        column_widths.get(col_idx, self.MIN_COLUMN_WIDTH),
                        current_width
                    )
                
                row_count += 1
            
            # Apply column widths
            for col_idx, width in column_widths.items():
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(self.MIN_COLUMN_WIDTH, width * 1.1)
            
            # Set header row height and freeze
            ws.row_dimensions[1].height = 25
            ws.freeze_panes = 'A2'
            
            # Save to buffer
            xlsx_buffer = BytesIO()
            wb.save(xlsx_buffer)
            xlsx_buffer.seek(0)
            
            # Get client name for filename
            client_name = ''
            if table.group and table.group.client:
                client_name = table.group.client.name

            # Generate filename
            filename = generate_export_filename(table.name, 'xlsx', client_name=client_name, status=status)
            
            # Create response — uses chunked streaming for large files
            xlsx_bytes = xlsx_buffer.getvalue()
            xlsx_buffer.close()
            response = stream_file_response(
                xlsx_bytes,
                filename,
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            
            return ExcelExportResult(
                success=True,
                response=response,
                filename=filename,
                row_count=row_count
            )
            
        except Exception as e:
            logger.error("Excel export failed: %s", e, exc_info=True)
            return ExcelExportResult(
                success=False,
                message='Excel export failed. Please try again or contact support.'
            )
    
    def _get_styles(self, Font, Alignment, Border, Side) -> Dict[str, Any]:
        """Get style definitions for Excel export."""
        return {
            'header_font': Font(name='Calibri', size=11, bold=True),
            'header_alignment': Alignment(horizontal='center', vertical='center'),
            'data_font': Font(name='Calibri', size=10),
            'data_alignment': Alignment(horizontal='left', vertical='center', wrap_text=False),
            'border': Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
        }


# =============================================================================
# MODULE-LEVEL CONVENIENCE FUNCTION
# =============================================================================

def export_cards_to_xlsx(table, cards: QuerySet, uppercase: bool = True) -> ExcelExportResult:
    """
    Convenience function to export cards to Excel.
    
    Args:
        table: IDCardTable instance
        cards: QuerySet of IDCard instances
        uppercase: Convert values to uppercase
        
    Returns:
        ExcelExportResult
    """
    exporter = ExcelExporter()
    return exporter.export_cards(table, cards, uppercase_values=uppercase)
