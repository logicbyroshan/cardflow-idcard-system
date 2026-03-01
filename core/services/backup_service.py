"""
Backup Service
==============

Background service that creates per-client ZIP archives containing:
  • One folder per IDCardTable
  • Inside each folder: XLSX files split by status (pending.xlsx, verified.xlsx …)
  • An ``images/`` folder holding all images referenced in the cards

Uses a daemon thread for the heavy I/O so the request returns immediately.
Progress is tracked in the BackupTask model, polled from the Manage Panel.

Auto-delete: sets ``auto_delete_at`` = now + 24 h after completion.
A lightweight cleanup thread checks periodically and deletes expired backups.
"""

import logging
import os
import shutil
import threading
import time
import zipfile
from datetime import timedelta
from io import BytesIO
from typing import List

from django.conf import settings
from django.utils import timezone

logger = logging.getLogger(__name__)

# Where backup ZIPs live (relative folder under MEDIA_ROOT)
BACKUP_DIR_NAME = os.path.join('temp', 'backups')
AUTO_DELETE_HOURS = 24
_cleanup_started = False
_cleanup_lock = threading.Lock()


def _backup_root():
    """Absolute path to the backup directory."""
    return os.path.join(settings.MEDIA_ROOT, BACKUP_DIR_NAME)


def _ensure_backup_dir(task_id: int):
    """Create per-task backup directory."""
    d = os.path.join(_backup_root(), str(task_id))
    os.makedirs(d, exist_ok=True)
    return d


# ─── Public API ───────────────────────────────────────────────────────────

def start_backup(task_id: int):
    """
    Launch the background thread that processes the backup.

    ``task_id`` must reference an existing ``BackupTask`` in *pending* state.
    """
    thread = threading.Thread(
        target=_process_backup,
        args=(task_id,),
        daemon=True,
    )
    thread.start()
    # Ensure the cleanup watcher is running
    _ensure_cleanup_thread()


def delete_backup_files(task_id: int):
    """Immediately delete files for a backup task and mark it deleted."""
    from core.models import BackupTask

    try:
        task = BackupTask.objects.get(pk=task_id)
        task.cleanup_files()
        task.status = 'deleted'
        task.save(update_fields=['status'])
        logger.info("Backup #%d files deleted.", task_id)
    except BackupTask.DoesNotExist:
        logger.warning("delete_backup_files: BackupTask #%d not found", task_id)


# ─── Background worker ───────────────────────────────────────────────────

def _process_backup(task_id: int):
    """Main background thread — iterates clients and builds ZIPs."""
    from core.models import BackupTask
    from client.models import Client
    from idcards.models import IDCardGroup, IDCardTable, IDCard

    try:
        task = BackupTask.objects.get(pk=task_id)
    except BackupTask.DoesNotExist:
        logger.error("BackupTask #%d vanished before processing", task_id)
        return

    task.status = 'processing'
    task.started_at = timezone.now()
    task.save(update_fields=['status', 'started_at'])

    out_dir = _ensure_backup_dir(task_id)

    try:
        client_ids: List[int] = task.client_ids or []
        clients = Client.objects.filter(pk__in=client_ids).order_by('name')
        task.total = clients.count()
        task.save(update_fields=['total'])

        zip_files = {}

        for idx, client in enumerate(clients, 1):
            task.current_client = client.name
            task.progress = idx - 1
            task.save(update_fields=['current_client', 'progress'])

            zip_info = _build_client_zip(client, out_dir)
            if zip_info:
                zip_files[str(client.pk)] = zip_info

        task.zip_files = zip_files
        task.progress = task.total
        task.current_client = ''
        task.status = 'completed'
        task.completed_at = timezone.now()
        task.auto_delete_at = timezone.now() + timedelta(hours=AUTO_DELETE_HOURS)
        task.save(update_fields=[
            'zip_files', 'progress', 'current_client',
            'status', 'completed_at', 'auto_delete_at',
        ])
        logger.info("Backup #%d completed — %d client ZIP(s)", task_id, len(zip_files))

    except Exception as exc:
        logger.exception("Backup #%d failed: %s", task_id, exc)
        try:
            task.refresh_from_db()
            task.status = 'failed'
            task.error_message = str(exc)[:2000]
            task.completed_at = timezone.now()
            task.save(update_fields=['status', 'error_message', 'completed_at'])
        except Exception:
            pass


def _build_client_zip(client, out_dir: str) -> dict | None:
    """
    Build a single client ZIP.

    Structure inside the ZIP::

        {ClientName}/
          {TableName}/
            pending.xlsx
            verified.xlsx
            …
            images/
              photo1.jpg
              photo2.jpg
    """
    from idcards.models import IDCardTable, IDCard
    from exports.utils import get_text_fields, get_image_fields, is_image_field

    groups = client.id_card_groups.all()
    if not groups.exists():
        return None

    date_str = timezone.now().strftime('%Y%m%d')
    safe_name = _safe_filename(client.name)
    zip_filename = f"{safe_name}_{date_str}.zip"
    zip_path = os.path.join(out_dir, zip_filename)

    try:
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            has_data = False
            for group in groups:
                tables = IDCardTable.objects.filter(group=group)
                for table in tables:
                    wrote = _write_table_to_zip(zf, safe_name, table)
                    if wrote:
                        has_data = True

        if not has_data:
            # Remove empty ZIP
            if os.path.exists(zip_path):
                os.remove(zip_path)
            return None

        file_size = os.path.getsize(zip_path)
        rel_path = os.path.relpath(zip_path, settings.MEDIA_ROOT).replace('\\', '/')
        return {
            'path': rel_path,
            'filename': zip_filename,
            'size': file_size,
        }

    except Exception as exc:
        logger.warning("Failed to build ZIP for client %s: %s", client.name, exc)
        return None


def _write_table_to_zip(zf: zipfile.ZipFile, client_folder: str, table) -> bool:
    """
    Write one table's data into the ZIP.

    Creates: ``{client}/{table}/status.xlsx`` + ``{client}/{table}/images/…``
    Returns True if at least one file was written.
    """
    from idcards.models import IDCard
    from exports.utils import get_text_fields, get_image_fields

    # Fetch all cards for this table
    all_cards = IDCard.objects.filter(table=table).order_by('id')
    if not all_cards.exists():
        return False

    table_safe = _safe_filename(table.name)
    base_prefix = f"{client_folder}/{table_safe}"

    text_fields = get_text_fields(table.fields or [])
    image_fields = get_image_fields(table.fields or [])

    # Group cards by status
    statuses = ['pending', 'verified', 'pool', 'approved', 'download', 'reprint']
    wrote_any = False
    image_paths_written = set()

    for status in statuses:
        cards = [c for c in all_cards if c.status == status]
        if not cards:
            continue

        # Build XLSX in memory
        xlsx_bytes = _build_xlsx_for_cards(cards, text_fields, image_fields, table.name)
        if xlsx_bytes:
            zf.writestr(f"{base_prefix}/{status}.xlsx", xlsx_bytes)
            wrote_any = True

        # Collect images
        for card in cards:
            _collect_images(zf, base_prefix, card, image_fields, image_paths_written)

    return wrote_any


def _build_xlsx_for_cards(cards, text_fields, image_fields, sheet_name: str) -> bytes | None:
    """Build an in-memory XLSX for a list of cards (single status group)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from exports.utils import format_field_value
    except ImportError:
        logger.error("openpyxl not installed — cannot build XLSX for backup")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # Styles
    hdr_font = Font(name='Arial', size=11, bold=True)
    hdr_align = Alignment(horizontal='center', vertical='center')
    data_font = Font(name='Arial', size=10)
    data_align = Alignment(horizontal='left', vertical='center')
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # Header
    headers = [f['name'] for f in text_fields] + [f['name'] for f in image_fields]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = border

    # Data rows
    text_count = len(text_fields)
    for ri, card in enumerate(cards, 2):
        fd = card.field_data or {}
        for ci, f in enumerate(text_fields, 1):
            val = format_field_value(fd.get(f['name'], ''), uppercase=True)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border
        for ii, img_f in enumerate(image_fields):
            ci = text_count + ii + 1
            raw = fd.get(img_f['name'], '')
            fname = _extract_image_stem(raw)
            cell = ws.cell(row=ri, column=ci, value=fname)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border

    # Auto-width
    for ci in range(1, len(headers) + 1):
        col_letter = get_column_letter(ci)
        max_len = max(
            (len(str(ws.cell(row=r, column=ci).value or '')) for r in range(1, len(cards) + 2)),
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(50, max(8, max_len + 2))
    ws.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _collect_images(zf, base_prefix, card, image_fields, already):
    """
    Add image files referenced by *card* into ``{base_prefix}/images/`` inside the ZIP.
    ``already`` is a set tracking arcnames written so far to avoid duplicates.
    """
    fd = card.field_data or {}
    for f in image_fields:
        raw = fd.get(f['name'], '')
        if not raw or raw in ('NOT_FOUND', ''):
            continue
        if raw.startswith('PENDING:'):
            continue

        # Determine the absolute path on disk
        rel = raw.replace('\\', '/')
        abs_path = os.path.join(settings.MEDIA_ROOT, rel)
        if not os.path.isfile(abs_path):
            continue

        arc_name = f"{base_prefix}/images/{os.path.basename(abs_path)}"
        if arc_name in already:
            continue
        already.add(arc_name)

        try:
            zf.write(abs_path, arc_name)
        except Exception as exc:
            logger.debug("Could not add image %s to ZIP: %s", abs_path, exc)


# ─── Cleanup watcher ─────────────────────────────────────────────────────

def _ensure_cleanup_thread():
    """Start a single daemon thread that auto-deletes expired backups."""
    global _cleanup_started
    with _cleanup_lock:
        if _cleanup_started:
            return
        _cleanup_started = True

    t = threading.Thread(target=_cleanup_loop, daemon=True)
    t.start()


def _cleanup_loop():
    """Periodically check for expired backups and delete them."""
    from core.models import BackupTask

    while True:
        try:
            now = timezone.now()
            expired = BackupTask.objects.filter(
                status='completed',
                is_auto_delete_cancelled=False,
                auto_delete_at__lte=now,
            )
            for task in expired:
                logger.info("Auto-deleting expired backup #%d", task.pk)
                task.cleanup_files()
                task.status = 'deleted'
                task.save(update_fields=['status'])
        except Exception as exc:
            logger.warning("Backup cleanup error: %s", exc)

        # Check every 5 minutes
        time.sleep(300)


# ─── Helpers ──────────────────────────────────────────────────────────────

def _safe_filename(name: str) -> str:
    """Convert a name into a filesystem-safe string."""
    import re
    safe = re.sub(r'[^\w\s-]', '', name).strip()
    safe = re.sub(r'[\s]+', '_', safe)
    return safe[:80] or 'unnamed'


def _extract_image_stem(raw: str) -> str:
    """Extract filename without extension from image path."""
    if not raw or raw in ('NOT_FOUND', ''):
        return ''
    val = str(raw).strip()
    if val.upper().startswith('PENDING:'):
        val = val[8:]
    basename = os.path.basename(val)
    name, _ = os.path.splitext(basename)
    return name
