"""
ID Card Bulk API — bulk upload, reupload images, and modals HTML.

Contains:
- api_idcard_bulk_upload
- _parse_excel_file, _parse_csv_file, _map_headers_to_fields
- api_idcard_reupload_images
- api_idcard_modals_html
"""
import json
import logging
import os

from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.conf import settings
from django.core.cache import cache as django_cache

from idcards.models import IDCard, IDCardTable
from ..services import IDCardService
from ..services.image_service import ImageService
from ..services.base import BaseService
from ..services.permission_service import (
    api_require_any_authenticated,
    api_require_permission,
)
from ..utils.upload_security import validate_zip_safety

from .idcard_helpers import (
    _safe_error,
    _check_client_scope_by_table,
    _CLIENT_READONLY_STATUSES,
    _is_client_readonly,
    validate_image_bytes,
)

# Logger for this module
logger = logging.getLogger(__name__)


@require_http_methods(["POST"])
@api_require_permission('perm_idcard_bulk_upload')
def api_idcard_bulk_upload(request, table_id):
    """API endpoint to bulk upload ID Cards from XLSX/CSV file with fuzzy matching and optional ZIP photo upload.
    
    Uses disk-based image storage for large ZIPs to prevent OOM.
    Small datasets (<50MB total) are kept in RAM for speed.
    Row processing is unified across XLSX and CSV via BulkUploadService.
    """
    _tbl, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    # Double-click guard: prevent duplicate uploads from rapid form submissions
    lock_key = f'bulk_upload_lock:{request.user.id}:{table_id}'
    if not django_cache.add(lock_key, 1, 300):
        return JsonResponse({'success': False, 'message': 'Upload already in progress. Please wait.'}, status=429)
    
    # Import the disk-backed image store and helpers
    from ..services.bulk_upload_service import (
        DiskBackedImageStore, extract_zip_to_store, process_data_rows,
        MAX_BULK_ROWS, BULK_BATCH_SIZE,
    )
    
    # Track all image stores for cleanup on exit
    _all_stores = []
    
    try:
        import openpyxl
        from io import BytesIO
        import re
        import zipfile
        import os
        import shutil
        from django.core.files.storage import default_storage
        from django.core.files.base import ContentFile

        # Pre-flight disk space check: require at least 500 MB free
        try:
            disk = shutil.disk_usage(settings.MEDIA_ROOT)
            if disk.free < 500 * 1024 * 1024:  # 500 MB
                return JsonResponse({
                    'success': False,
                    'message': 'Insufficient disk space. Please contact your administrator.'
                }, status=507)
        except Exception:
            pass  # Non-critical — proceed if check fails
        
        table = get_object_or_404(IDCardTable.objects.select_related('group__client'), id=table_id)
        
        if 'file' not in request.FILES:
            return JsonResponse({'success': False, 'message': 'No file uploaded!'}, status=400)
        
        uploaded_file = request.FILES['file']
        file_name = uploaded_file.name.lower()
        file_size = uploaded_file.size
        
        # Get image field names from table using BaseService
        image_field_names = BaseService.get_image_field_names(table.fields)
        
        # ── Extract ZIP images into disk-backed stores ──
        # Each image field gets its own DiskBackedImageStore
        zip_photos_by_field = {}  # { field_name: DiskBackedImageStore }
        
        # Check for multiple ZIP files - one per image field
        zip_field_names_str = request.POST.get('zip_field_names', '[]')
        try:
            zip_field_names = json.loads(zip_field_names_str)
        except (json.JSONDecodeError, TypeError):
            zip_field_names = []
        
        logger.debug("zip_field_names = %s", zip_field_names)
        
        # Process per-field ZIP files
        for field_name in zip_field_names:
            zip_key = f'photos_zip_{field_name}'
            if zip_key in request.FILES:
                photos_zip_file = request.FILES[zip_key]
                store = DiskBackedImageStore()
                _all_stores.append(store)
                count = extract_zip_to_store(photos_zip_file, store)
                if count > 0:
                    zip_photos_by_field[field_name] = store
                    logger.debug("Field '%s' extracted %d images", field_name, count)
        
        # Legacy: single photos_zip (backward compatibility)
        if not zip_photos_by_field and 'photos_zip' in request.FILES:
            photos_zip_file = request.FILES['photos_zip']
            first_image_field = image_field_names[0] if image_field_names else 'PHOTO'
            store = DiskBackedImageStore()
            _all_stores.append(store)
            count = extract_zip_to_store(photos_zip_file, store)
            if count > 0:
                zip_photos_by_field[first_image_field] = store
        
        # Unified ZIP files (images auto-matched to all columns)
        unified_zip_photos = DiskBackedImageStore()
        _all_stores.append(unified_zip_photos)
        
        try:
            unified_zip_count = min(int(request.POST.get('unified_zip_count', 0)), 20)
        except (ValueError, TypeError):
            unified_zip_count = 0
        
        for i in range(unified_zip_count):
            zip_key = f'unified_zip_{i}'
            if zip_key in request.FILES:
                extract_zip_to_store(request.FILES[zip_key], unified_zip_photos)
        
        logger.debug("unified_zip_photos count = %d", len(unified_zip_photos))
        
        # Get client for ImageService operations
        client = table.group.client
        
        # Get all table fields
        all_table_fields = table.fields
        table_fields = [f['name'] for f in all_table_fields if not BaseService.is_image_field(f)]
        image_fields = [f['name'] for f in all_table_fields if BaseService.is_image_field(f)]
        
        matched_field_names = []
        
        # Check if frontend sent a manual field mapping
        frontend_mapping_str = request.POST.get('field_mapping', '')
        frontend_mapping = {}
        if frontend_mapping_str:
            try:
                frontend_mapping = json.loads(frontend_mapping_str)
                if not isinstance(frontend_mapping, dict):
                    frontend_mapping = {}
            except (json.JSONDecodeError, TypeError):
                frontend_mapping = {}
        
        # ── Parse file (XLSX/XLS/CSV) ──
        if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
            rows_data, headers, header_to_field, image_ref_columns, matched_field_names, parse_error = \
                _parse_excel_file(uploaded_file, file_name, table_fields, image_fields,
                                  frontend_mapping, all_table_fields)
            if parse_error:
                return parse_error
            is_csv = False
        elif file_name.endswith('.csv'):
            rows_data, headers, header_to_field, image_ref_columns, matched_field_names, parse_error = \
                _parse_csv_file(uploaded_file, table_fields, image_fields,
                                frontend_mapping, all_table_fields)
            if parse_error:
                return parse_error
            is_csv = True
        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid file format! Please upload .xlsx, .xls, or .csv file.'
            }, status=400)
        
        if not header_to_field:
            return JsonResponse({
                'success': False,
                'message': f'No matching columns found! Expected columns: {", ".join(table_fields)}'
            }, status=400)
        
        if len(rows_data) > MAX_BULK_ROWS:
            return JsonResponse({
                'success': False,
                'message': f'File has {len(rows_data)} rows. Maximum allowed is {MAX_BULK_ROWS}.'
            }, status=400)
        
        # Reverse rows so first Excel row gets highest DB id (preserves order in -id display)
        rows_data = list(reversed(rows_data))
        
        # ── Process rows using unified service ──
        result = process_data_rows(
            rows=rows_data,
            header_to_field=header_to_field,
            image_ref_columns=image_ref_columns,
            image_fields=image_fields,
            all_table_fields=all_table_fields,
            table=table,
            client=client,
            zip_photos_by_field=zip_photos_by_field,
            unified_zip_photos=unified_zip_photos,
            request_user=request.user,
            is_csv=is_csv,
        )
        
        cards_created = result['cards_created']
        total_photos_matched = result['total_photos_matched']
        errors = result['errors']
        
        # Return result
        photo_msg = f" with {total_photos_matched} photos matched" if total_photos_matched > 0 else ""
        response = {
            'success': True,
            'message': f'Successfully created {cards_created} ID cards{photo_msg}!',
            'cards_created': cards_created,
            'photos_matched': total_photos_matched,
            'matched_fields': matched_field_names,
        }
        
        if errors:
            response['errors'] = errors[:10]
            response['error_count'] = len(errors)
        
        return JsonResponse(response)
        
    except ImportError:
        return JsonResponse({
            'success': False,
            'message': 'openpyxl library not installed. Run: pip install openpyxl'
        }, status=500)
    except Exception as e:
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)
    finally:
        # Always cleanup disk-backed image stores
        for store in _all_stores:
            try:
                store.cleanup()
            except Exception:
                pass
        django_cache.delete(lock_key)


def _parse_excel_file(uploaded_file, file_name, table_fields, image_fields,
                       frontend_mapping, all_table_fields):
    """Parse Excel file and return (rows_data, headers, header_to_field, image_ref_columns, matched_fields, error_response).
    Returns error_response=None on success, or a JsonResponse on failure."""
    import openpyxl
    from io import BytesIO
    
    try:
        file_content = uploaded_file.read()
        if len(file_content) < 4:
            return None, None, None, None, [], JsonResponse({
                'success': False, 'message': 'File is too small or empty.'
            }, status=400)
        
        magic_bytes = file_content[:4]
        is_zip = magic_bytes[:2] == b'PK'
        is_old_xls = (magic_bytes[0] == 0xD0 and magic_bytes[1] == 0xCF)
        
        headers = []
        rows_data = []
        
        if is_zip or file_name.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(BytesIO(file_content))
                ws = wb.active
                for cell in ws[1]:
                    if cell.value:
                        headers.append(str(cell.value).strip()
                                       .replace('_x000D_', '').replace('_X000D_', '')
                                       .replace('_x000d_', '').replace('\r', ''))
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows_data.append(row)
            except Exception as xlsx_error:
                if not is_zip:
                    is_old_xls = True
                else:
                    raise xlsx_error
        
        if is_old_xls or (file_name.endswith('.xls') and not file_name.endswith('.xlsx') and not headers):
            try:
                import xlrd
                wb = xlrd.open_workbook(file_contents=file_content)
                ws = wb.sheet_by_index(0)
                headers = []
                for col_idx in range(ws.ncols):
                    cell_value = ws.cell_value(0, col_idx)
                    if cell_value:
                        headers.append(str(cell_value).strip())
                rows_data = []
                for row_idx in range(1, ws.nrows):
                    row = []
                    for col_idx in range(ws.ncols):
                        row.append(ws.cell_value(row_idx, col_idx))
                    rows_data.append(tuple(row))
            except ImportError:
                return None, None, None, None, [], JsonResponse({
                    'success': False, 'message': 'xlrd library not installed for .xls files.'
                }, status=400)
            except Exception:
                return None, None, None, None, [], JsonResponse({
                    'success': False, 'message': 'Error reading .xls file. Please check the file format.'
                }, status=400)
        
        if not headers:
            return None, None, None, None, [], JsonResponse({
                'success': False, 'message': 'Could not read headers from Excel file.'
            }, status=400)
    except Exception:
        return None, None, None, None, [], JsonResponse({
            'success': False, 'message': 'Error reading Excel file. Please check the file format.'
        }, status=400)
    
    header_to_field, image_ref_columns, matched_field_names = _map_headers_to_fields(
        headers, table_fields, image_fields, frontend_mapping, all_table_fields, is_csv=False
    )
    
    return rows_data, headers, header_to_field, image_ref_columns, matched_field_names, None


def _parse_csv_file(uploaded_file, table_fields, image_fields,
                     frontend_mapping, all_table_fields):
    """Parse CSV file and return (rows_data, headers, header_to_field, image_ref_columns, matched_fields, error_response)."""
    import csv
    from io import StringIO
    
    try:
        content = uploaded_file.read().decode('utf-8-sig')
        reader = csv.DictReader(StringIO(content))
        csv_headers = reader.fieldnames or []
        rows_data = list(reader)
    except Exception:
        return None, None, None, None, [], JsonResponse({
            'success': False, 'message': 'Error reading CSV file. Please check the file format.'
        }, status=400)
    
    if not csv_headers:
        return None, None, None, None, [], JsonResponse({
            'success': False, 'message': 'Could not read headers from CSV file.'
        }, status=400)
    
    header_to_field, image_ref_columns, matched_field_names = _map_headers_to_fields(
        csv_headers, table_fields, image_fields, frontend_mapping, all_table_fields, is_csv=True
    )
    
    return rows_data, csv_headers, header_to_field, image_ref_columns, matched_field_names, None


def _map_headers_to_fields(headers, table_fields, image_fields, frontend_mapping,
                            all_table_fields, *, is_csv=False):
    """Map file headers to table fields using fuzzy matching or frontend mapping.
    Returns (header_to_field, image_ref_columns, matched_field_names)."""
    header_to_field = {}
    available_fields = table_fields.copy()
    image_ref_columns = {}
    unmatched_image_fields = list(image_fields)
    matched_field_names = []
    
    if frontend_mapping:
        if is_csv:
            for table_field_name, excel_header in frontend_mapping.items():
                if table_field_name in available_fields and excel_header in headers:
                    header_to_field[excel_header] = table_field_name
                    available_fields.remove(table_field_name)
                    matched_field_names.append(table_field_name)
        else:
            header_index = {h: i for i, h in enumerate(headers)}
            for table_field_name, excel_header in frontend_mapping.items():
                if table_field_name in available_fields and excel_header in header_index:
                    idx = header_index[excel_header]
                    header_to_field[idx] = table_field_name
                    available_fields.remove(table_field_name)
                    matched_field_names.append(table_field_name)
        
        # Auto-match image columns
        for idx_or_header in (range(len(headers)) if not is_csv else headers):
            header = headers[idx_or_header] if not is_csv else idx_or_header
            if not header:
                continue
            if not is_csv and idx_or_header in header_to_field:
                continue
            if is_csv and header in header_to_field:
                continue
            matched_img = BaseService.find_best_image_field_match(header, unmatched_image_fields)
            if matched_img:
                image_ref_columns[matched_img] = idx_or_header if not is_csv else header
                unmatched_image_fields.remove(matched_img)
    else:
        # Auto fuzzy matching
        for idx_or_header in (range(len(headers)) if not is_csv else headers):
            header = headers[idx_or_header] if not is_csv else idx_or_header
            if not header:
                continue
            
            matched_img = BaseService.find_best_image_field_match(header, unmatched_image_fields)
            if matched_img:
                image_ref_columns[matched_img] = idx_or_header if not is_csv else header
                unmatched_image_fields.remove(matched_img)
                continue
            
            header_str = header if is_csv else header
            match = BaseService.find_best_field_match(header_str.strip() if is_csv else header, available_fields)
            if match:
                if is_csv:
                    header_to_field[header] = match
                else:
                    header_to_field[idx_or_header] = match
                available_fields.remove(match)
                matched_field_names.append(match)
    
    return header_to_field, image_ref_columns, matched_field_names


@require_http_methods(["POST"])
@api_require_permission('perm_idcard_bulk_reupload')
def api_idcard_reupload_images(request, table_id):
    """
    API endpoint to reupload images from a ZIP file.
    Matches ZIP filenames to card image references (PENDING: or existing paths) and updates them.
    
    Supports:
    - PENDING:reference matching (for cards created without images)
    - Existing image path updates (applies edit naming: original_14 + _HHMMSS)
    - Multiple image fields per card
    - Thumbnail generation for all saved images
    """
    _tbl, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    # Client/client_staff cannot reupload images for tables with approved/download/reprint cards
    if request.user.role in ('client', 'client_staff'):
        has_locked = IDCard.objects.filter(
            table_id=table_id, status__in=_CLIENT_READONLY_STATUSES
        ).exists()
        if has_locked:
            return JsonResponse({
                'success': False,
                'message': 'This table contains cards in approved/download status. Client users cannot reupload images.'
            }, status=403)
    # Double-click guard: prevent duplicate reupload from rapid form submissions
    lock_key = f'reupload_lock:{request.user.id}:{table_id}'
    if not django_cache.add(lock_key, 1, 300):
        return JsonResponse({'success': False, 'message': 'Reupload already in progress. Please wait.'}, status=429)
    try:
        import zipfile
        from django.db import transaction
        
        table = get_object_or_404(IDCardTable.objects.select_related('group__client'), id=table_id)
        client = table.group.client
        
        if 'photos_zip' not in request.FILES:
            return JsonResponse({'success': False, 'message': 'No ZIP file uploaded!'}, status=400)
        
        # Get image field names from table
        image_field_names = BaseService.get_image_field_names(table.fields)
        if not image_field_names:
            return JsonResponse({'success': False, 'message': 'No image fields defined in table!'}, status=400)
        
        # Get target field from request (optional - defaults to first image field)
        target_field = request.POST.get('target_field', image_field_names[0])
        if target_field not in image_field_names:
            target_field = image_field_names[0]
        
        # Extract photos from ZIP — use temp file if available (avoids OOM on large uploads)
        zip_photos = {}  # { normalized_key: { bytes, ext, original_name } }
        
        try:
            zip_file = request.FILES['photos_zip']

            # ZIP size guard
            if hasattr(zip_file, 'size') and zip_file.size > 600 * 1024 * 1024:
                return JsonResponse({'success': False, 'message': 'ZIP file exceeds 600 MB limit.'}, status=400)
            
            # ZIP bomb / nested archive check
            zok, zerr = validate_zip_safety(zip_file)
            if not zok:
                return JsonResponse({'success': False, 'message': zerr}, status=400)

            # Open ZIP directly from file handle (Django spills >10MB to /tmp)
            if hasattr(zip_file, 'temporary_file_path'):
                zf = zipfile.ZipFile(zip_file.temporary_file_path(), 'r')
            else:
                zip_file.seek(0)
                zf = zipfile.ZipFile(zip_file, 'r')

            with zf:
                for zip_info in zf.infolist():
                    if zip_info.is_dir():
                        continue
                    
                    file_in_zip = zip_info.filename
                    base_name = os.path.basename(file_in_zip)
                    name_without_ext = os.path.splitext(base_name)[0]
                    ext = os.path.splitext(base_name)[1].lower()
                    
                    if ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']:
                        try:
                            image_bytes = zf.read(zip_info.filename)
                            is_valid, error_msg = validate_image_bytes(image_bytes)
                            if is_valid:
                                normalized_key = BaseService.normalize_image_identifier(name_without_ext)
                                if normalized_key:
                                    # Deterministic: if duplicate key, keep alphabetically-first filename
                                    existing = zip_photos.get(normalized_key)
                                    if existing is None or base_name < existing['original_name']:
                                        zip_photos[normalized_key] = {
                                            'bytes': image_bytes,
                                            'ext': ext,
                                            'original_name': base_name
                                        }
                        except Exception:
                            continue
        except Exception as zip_error:
            logger.exception('Error reading ZIP file: %s', zip_error)
            return JsonResponse({'success': False, 'message': 'Error reading ZIP file. Please check the file and try again.'}, status=400)
        
        if not zip_photos:
            return JsonResponse({'success': False, 'message': 'No valid images found in ZIP file!'}, status=400)
        
        logger.debug("Reupload: %d images extracted from ZIP, keys: %s", len(zip_photos), list(zip_photos.keys())[:10])
        
        # Get cards — scoped to selected IDs if provided, else all in table for current status
        card_ids = []
        if 'card_ids' in request.POST:
            try:
                card_ids = json.loads(request.POST.get('card_ids', '[]'))
            except (json.JSONDecodeError, TypeError):
                card_ids = []
        
        # Filter out empty/falsy values
        card_ids = [int(cid) for cid in card_ids if cid and str(cid).strip().isdigit()] if card_ids else []
        
        if card_ids:
            cards_qs = IDCard.objects.filter(table=table, id__in=card_ids).order_by('id')
        else:
            # No specific IDs — reupload to ALL cards in this table (filtered by status if provided)
            status_filter = request.POST.get('status', '')
            if status_filter and status_filter in BaseService.VALID_STATUSES:
                cards_qs = IDCard.objects.filter(table=table, status=status_filter).order_by('id')
            else:
                cards_qs = IDCard.objects.filter(table=table).order_by('id')
        
        updated_count = 0
        matched_count = 0
        errors = []
        
        # Process cards in batches to avoid long-running transactions
        # (SQLite locks the entire DB during a transaction; smaller batches
        # reduce lock duration and prevent "database is locked" errors)
        REUPLOAD_BATCH_SIZE = 50
        all_card_ids = list(cards_qs.values_list('id', flat=True))
        batch_counter = 0
        
        for batch_start in range(0, len(all_card_ids), REUPLOAD_BATCH_SIZE):
            batch_ids = all_card_ids[batch_start:batch_start + REUPLOAD_BATCH_SIZE]
            
            with transaction.atomic():
                batch_cards = IDCard.objects.filter(id__in=batch_ids).order_by('id')
                
                for card in batch_cards:
                    field_data = card.field_data or {}
                    card_updated = False
                    
                    for img_field in image_field_names:
                        current_value = field_data.get(img_field, '')
                        
                        # Determine what to match against
                        match_key = None
                        existing_path = None
                        
                        if current_value.startswith('PENDING:'):
                            # Extract the reference from PENDING:reference
                            match_key = BaseService.normalize_image_identifier(current_value[8:])
                        elif current_value and current_value not in ('NOT_FOUND', ''):
                            # Has existing image - extract filename for matching
                            existing_path = current_value
                            existing_filename = os.path.splitext(os.path.basename(current_value))[0]
                            match_key = BaseService.normalize_image_identifier(existing_filename)
                        else:
                            # No current value - skip unless we want to match by card data
                            # Could extend to match by NAME or other field values
                            continue
                        
                        if not match_key:
                            continue
                        
                        # Try to find matching photo in ZIP
                        if match_key in zip_photos:
                            photo_info = zip_photos[match_key]
                            matched_count += 1
                            
                            try:
                                batch_counter += 1
                                
                                # Use single-authority entry point
                                if existing_path:
                                    result = ImageService.replace_image(
                                        image_bytes=photo_info['bytes'],
                                        client=client,
                                        field_name=img_field,
                                        existing_path=existing_path,
                                        card=card,
                                        batch_counter=batch_counter,
                                        original_ext=photo_info['ext'],
                                    )
                                else:
                                    result = ImageService.save_new_image(
                                        image_bytes=photo_info['bytes'],
                                        client=client,
                                        field_name=img_field,
                                        card=card,
                                        batch_counter=batch_counter,
                                        original_ext=photo_info['ext'],
                                    )
                                
                                if result.success and result.data.get('final_value'):
                                    field_data[img_field] = result.data['final_value']
                                    card_updated = True
                                    logger.debug("Reupload: Card %s field %s updated to %s", 
                                               card.pk, img_field, result.data['final_value'])
                                else:
                                    errors.append(f"Card {card.pk}: Failed to save {img_field} - {result.message}")
                            except Exception as save_err:
                                errors.append(f"Card {card.pk}: Error saving {img_field} - {str(save_err)}")
                    
                    if card_updated:
                        card.field_data = field_data
                        card.save()
                        updated_count += 1
        
        # Build response
        result_msg = f"Updated {updated_count} cards with {matched_count} images matched"
        response = {
            'success': True,
            'message': result_msg,
            'updated_count': updated_count,
            'matched_count': matched_count,
            'zip_images_count': len(zip_photos),
        }
        
        if errors:
            response['errors'] = errors[:10]
            response['error_count'] = len(errors)
        
        return JsonResponse(response)
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)
    finally:
        django_cache.delete(lock_key)


# ==================== MODALS HTML (Lazy Load) ====================

@require_http_methods(["GET"])
@api_require_any_authenticated
def api_idcard_modals_html(request, table_id):
    """Return rendered modals.html partial for lazy-loading.
    Used by modal-loader.js to inject modals on first user interaction
    instead of pre-rendering them on every page load.
    """
    from django.template.loader import render_to_string
    from django.http import HttpResponse

    table, err = _check_client_scope_by_table(request.user, table_id)
    if err:
        return err

    html = render_to_string('partials/idcard/modals.html', {'table': table}, request=request)
    return HttpResponse(html, content_type='text/html; charset=utf-8')
