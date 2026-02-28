"""
ID Card API Views
Contains: All ID Card Table and ID Card related API endpoints.
Including: CRUD, bulk operations, search, status changes, bulk upload.

ARCHITECTURE RULES (enforced):
- Views are ULTRA-THIN: parse request → call service → return JsonResponse.
- NO .save(), .create(), .delete() on IDCard / IDCardTable in views.
- All mutations delegate to IDCardService or WorkflowService.
- Permission enforcement via @api_require_permission decorators.
- Client scoping via _check_client_scope_by_table / _check_client_scope_by_card.

EXCEPTION: api_idcard_bulk_upload and api_idcard_reupload_images still
contain inline mutation logic — flagged for future service extraction.
"""
import json
import logging
import os

from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.conf import settings
from django.core.cache import cache as django_cache

from ..models import IDCardGroup, IDCard, IDCardTable
from ..services import IDCardService
from ..services.image_service import ImageService
from ..services.base import BaseService
from ..services.activity_service import ActivityService
from ..services.permission_service import (
    PermissionService,
    api_require_any_authenticated,
    api_require_permission,
)
from ..services.workflow_service import WorkflowService
from ..utils.upload_security import validate_zip_safety

# Logger for this module
logger = logging.getLogger(__name__)


def _safe_error(e, fallback='An error occurred. Please try again.'):
    """Return a safe error message for API responses. Logs the real exception."""
    logger.exception("API error: %s", e)
    return fallback


def _build_class_filter_q(qs, class_filter, class_field_name):
    """Apply class filter with canonical normalization.

    Finds ALL raw variants in the DB that normalize to the same canonical
    value as the filter, then matches them with __in.  This ensures
    'KG-I', 'KG1', 'KGI', 'LKG' are all captured when filtering by 'KG1'.
    """
    from django.db.models.fields.json import KeyTextTransform
    from django.db.models.functions import Cast
    from django.db.models import CharField, Q
    from core.utils.field_utils import normalize_class_value

    norm_filter = normalize_class_value(class_filter)

    # Get all distinct raw class values in the queryset
    all_raw = list(
        qs.annotate(_cv_raw=Cast(KeyTextTransform(class_field_name, 'field_data'), CharField()))
        .exclude(_cv_raw__isnull=True).exclude(_cv_raw='')
        .order_by()
        .values_list('_cv_raw', flat=True).distinct()
    )

    # Find raw values that normalize to the same canonical
    matching_raw = [r for r in all_raw if normalize_class_value(r) == norm_filter]

    if not matching_raw:
        return qs.none()

    # Build filter: match any of the raw variants
    qs = qs.annotate(_cls=KeyTextTransform(class_field_name, 'field_data'))
    q = Q()
    for raw in matching_raw:
        q |= Q(_cls=raw)
    return qs.filter(q)


def _get_class_section_field_names(table):
    """Extract class and section field names from a table's field definitions.

    Matches by type OR by name (mirrors IDCardTable.has_class_field / has_section_field).
    Returns (class_field_name, section_field_name) — either may be None.
    """
    class_field = None
    section_field = None
    for field in (table.fields or []):
        ftype = field.get('type', '')
        fname = field.get('name', '')
        fname_lower = fname.lower() if fname else ''
        if not class_field and (ftype == 'class' or fname_lower == 'class'):
            class_field = fname
        elif not section_field and (ftype == 'section' or fname_lower == 'section'):
            section_field = fname
    return class_field, section_field


# ==================== ADMIN STAFF CLIENT SCOPING ====================
# Ensures admin_staff can only access data belonging to their assigned clients.

def _access_denied_response():
    """Factory: return a fresh 403 JsonResponse per request (thread-safe)."""
    return JsonResponse(
        {'success': False, 'message': 'Access denied. You are not assigned to this client.'},
        status=403,
    )

def _check_client_scope_by_group(user, group_id):
    """Check user has access to the client owning this group. Returns (group, error_response).
    
    Delegates to PermissionService.can_access_client() (single authority).
    """
    group = get_object_or_404(IDCardGroup, id=group_id)
    if not PermissionService.can_access_client(user, group.client_id):
        return None, _access_denied_response()
    return group, None

def _check_client_scope_by_table(user, table_id):
    """Check user has access to the client owning this table. Returns (table, error_response).
    
    Delegates to PermissionService.can_access_client() (single authority).
    """
    table = get_object_or_404(IDCardTable.objects.select_related('group'), id=table_id)
    if not PermissionService.can_access_client(user, table.group.client_id):
        return None, _access_denied_response()
    return table, None

def _check_client_scope_by_card(user, card_id):
    """Check user has access to the client owning this card. Returns (card, error_response).
    
    Delegates to PermissionService.can_access_client() (single authority).
    """
    card = get_object_or_404(IDCard.objects.select_related('table__group'), id=card_id)
    if not PermissionService.can_access_client(user, card.table.group.client_id):
        return None, _access_denied_response()
    return card, None


# ==================== CLIENT READONLY ON APPROVED+ ====================
# After cards reach approved/download/reprint, client & client_staff users
# can only VIEW — no edit, delete, status change, or image reupload.

_CLIENT_READONLY_STATUSES = frozenset({'approved', 'download', 'reprint'})

def _client_readonly_response():
    """Fresh 403 response for each request."""
    return JsonResponse(
        {'success': False, 'message': 'Cards in approved / download status cannot be modified by client users.'},
        status=403,
    )

def _is_client_readonly(user, card_status):
    """Return True when client/client_staff tries to modify a card in a locked status."""
    return user.role in ('client', 'client_staff') and card_status in _CLIENT_READONLY_STATUSES


# ==================== FIELD HELPERS (canonical: core.utils.field_utils) ====================
# Re-exported for backward compatibility within this view module.
# All new code should import directly from core.utils.field_utils.
from core.utils.field_utils import (
    validate_image_bytes,
    convert_class_value,
    convert_section_value,
    NUMERIC_TO_ROMAN,
    VALID_CLASS_VALUES,
    CLASS_UPGRADE_MAP,
)


# ==================== ID CARD TABLE API ENDPOINTS ====================

@require_http_methods(["POST"])
@api_require_permission('perm_idcard_setting_add')
def api_idcard_table_create(request, group_id):
    """API endpoint to create a new ID Card Table"""
    group, err = _check_client_scope_by_group(request.user, group_id)
    if err: return err
    try:
        data = json.loads(request.body)
        result = IDCardService.create_table(group_id, data)
        return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data!'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["GET"])
@api_require_permission('perm_idcard_setting_list')
def api_idcard_table_get(request, table_id):
    """API endpoint to get a single ID Card Table"""
    table, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    result = IDCardService.get_table(table_id)
    return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)


@require_http_methods(["POST", "PUT"])
@api_require_permission('perm_idcard_setting_edit')
def api_idcard_table_update(request, table_id):
    """API endpoint to update an ID Card Table"""
    table, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    try:
        data = json.loads(request.body)
        result = IDCardService.update_table(table_id, data)
        return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data!'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["DELETE", "POST"])
@api_require_permission('perm_idcard_setting_delete')
def api_idcard_table_delete(request, table_id):
    """API endpoint to delete an ID Card Table"""
    table, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    try:
        result = IDCardService.delete_table(table_id)
        return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)
    except Exception as e:
        logger.exception("Table delete error: %s", e)
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["POST"])
@api_require_permission('perm_idcard_setting_status')
def api_idcard_table_toggle_status(request, table_id):
    """API endpoint to toggle ID Card Table active/inactive status"""
    table, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    try:
        result = IDCardService.toggle_table_status(table_id)
        return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)
    except Exception as e:
        logger.exception("Table toggle status error: %s", e)
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["GET"])
@api_require_permission('perm_idcard_setting_list')
def api_idcard_table_list(request, group_id):
    """API endpoint to list all ID Card Tables for a group"""
    group, err = _check_client_scope_by_group(request.user, group_id)
    if err: return err
    result = IDCardService.list_tables(group_id)
    return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)


# ==================== CREATE TABLE FROM XLSX ====================

# Field-type inference patterns: map XLSX header names to field types.
# Order matters: first match wins.  All comparisons are case-insensitive.
_HEADER_TYPE_MAP = [
    # Image-like fields
    (['mother photo', 'm photo', 'mother_photo', 'mother pic'], 'mother_photo'),
    (['father photo', 'f photo', 'father_photo', 'father pic'], 'father_photo'),
    (['photo', 'pic', 'picture', 'image', 'student photo', 'student image'], 'photo'),
    (['signature', 'sign'], 'signature'),
    (['barcode'], 'barcode'),
    (['qr code', 'qr_code', 'qr'], 'qr_code'),
    # Structural fields
    (['class'], 'class'),
    (['section', 'sec'], 'section'),
    (['email', 'e-mail', 'email id', 'email address'], 'email'),
]


def _infer_field_type(header_name: str) -> str:
    """Infer field type from an XLSX header name.

    Returns one of the VALID_FIELD_TYPES for IDCardTable.
    Falls back to 'text' for any unrecognised header.
    """
    normalized = header_name.strip().lower().replace('_', ' ')
    for patterns, field_type in _HEADER_TYPE_MAP:
        if normalized in patterns:
            return field_type
    return 'text'


@require_http_methods(["POST"])
@api_require_permission('perm_idcard_setting_add')
def api_create_table_from_xlsx(request, group_id):
    """
    Create a new IDCardTable from an XLSX file's header row, then bulk-upload
    the data rows into the table.  Optionally accepts ZIP files for image fields.

    This combines two steps into one:
      1. Reads the first row of the XLSX to derive field names + types.
      2. Creates a table with those fields.
      3. Delegates to the existing bulk-upload logic to import the data.

    POST /api/group/<group_id>/table/create-from-xlsx/

    Form-data:
        file              : XLSX/XLS/CSV file   (required)
        table_name        : Optional override for the table name
        photos_zip_<FIELD>: ZIP per image field  (optional)
        unified_zip_<N>   : Unified ZIPs         (optional)
        unified_zip_count : Number of unified ZIPs (optional)
        zip_field_names   : JSON array of field names with ZIP uploads (optional)

    Returns JSON:
        { success, message, table_id, table_name, cards_created, ... }
    """
    import openpyxl
    from io import BytesIO

    group, err = _check_client_scope_by_group(request.user, group_id)
    if err:
        return err

    # ── 1. Validate file ────────────────────────────────────────────
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'No file uploaded.'}, status=400)

    uploaded_file = request.FILES['file']
    file_name = uploaded_file.name.lower()
    if not file_name.endswith(('.xlsx', '.xls', '.csv')):
        return JsonResponse({
            'success': False,
            'level': 'warning',
            'message': 'Only .xlsx, .xls, and .csv files are supported.'
        }, status=400)

    # Size guard: 50 MB max for the spreadsheet
    if uploaded_file.size > 50 * 1024 * 1024:
        return JsonResponse({
            'success': False,
            'level': 'warning',
            'message': 'Spreadsheet file must be under 50 MB.'
        }, status=400)

    # ── 2. Read headers ─────────────────────────────────────────────
    try:
        if file_name.endswith('.csv'):
            import csv, io
            content = uploaded_file.read().decode('utf-8-sig', errors='replace')
            uploaded_file.seek(0)
            reader = csv.reader(io.StringIO(content))
            headers = next(reader, [])
        else:
            wb = openpyxl.load_workbook(BytesIO(uploaded_file.read()), read_only=True, data_only=True)
            uploaded_file.seek(0)
            ws = wb.active
            raw_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            headers = [
                str(cell).strip().replace('_x000D_', '').replace('_X000D_', '').replace('_x000d_', '').replace('\r', '')
                if cell is not None else ''
                for cell in raw_row
            ]
            wb.close()
    except Exception as exc:
        logger.error("Failed to read XLSX headers: %s", exc)
        return JsonResponse({
            'success': False,
            'message': 'Could not read the spreadsheet. Please check the file format.'
        }, status=400)

    # Filter out empty headers
    headers = [h for h in headers if h]
    if not headers:
        return JsonResponse({
            'success': False, 'message': 'The spreadsheet has no column headers in the first row.'
        }, status=400)

    if len(headers) > IDCardService.MAX_FIELDS_PER_TABLE:
        return JsonResponse({
            'success': False,
            'level': 'warning',
            'message': f'Maximum {IDCardService.MAX_FIELDS_PER_TABLE} columns allowed. '
                       f'Your file has {len(headers)} columns.'
        }, status=400)

    # ── 3. Infer field definitions (or use client-provided config) ──
    field_config_json = request.POST.get('field_config', '')
    client_field_config = None
    if field_config_json:
        try:
            import json as _json
            client_field_config = _json.loads(field_config_json)
            if not isinstance(client_field_config, list):
                client_field_config = None
        except (ValueError, TypeError):
            client_field_config = None

    VALID_FIELD_TYPES = {
        'text', 'class', 'section', 'email', 'photo',
        'mother_photo', 'father_photo', 'signature', 'barcode', 'qr_code',
    }

    fields = []
    for idx, header in enumerate(headers):
        # Default: auto-infer
        field_type = _infer_field_type(header)
        mandatory = False

        # Override with client-provided config if available and valid
        if client_field_config and idx < len(client_field_config):
            cfg = client_field_config[idx]
            if isinstance(cfg, dict):
                cfg_type = cfg.get('type', '')
                if cfg_type in VALID_FIELD_TYPES:
                    field_type = cfg_type
                mandatory = bool(cfg.get('mandatory', False))

        fields.append({
            'name': header.strip().upper(),
            'type': field_type,
            'order': idx,
            'mandatory': mandatory,
        })

    # ── 4. Create the table ─────────────────────────────────────────
    table_name = (request.POST.get('table_name') or '').strip().upper()
    if not table_name:
        # Derive from filename (strip extension)
        import os as _os
        base = _os.path.splitext(uploaded_file.name)[0]
        table_name = base.strip().upper()[:255] or 'IMPORTED TABLE'

    try:
        table = IDCardTable.objects.create(
            group=group,
            name=table_name,
            fields=fields,
            is_active=True,
        )
    except Exception as exc:
        logger.exception("Failed to create table from XLSX: %s", exc)
        return JsonResponse({
            'success': False, 'message': 'Failed to create table. Please try again.'
        }, status=500)

    logger.info(
        "Created table %d (%s) with %d fields from XLSX (user=%s)",
        table.id, table_name, len(fields), request.user.id,
    )

    # ── 5. Delegate to existing bulk upload ─────────────────────────
    # We re-use the synchronous bulk upload by faking the table_id into the
    # existing function.  The uploaded file + ZIPs are already in request.FILES.
    try:
        # Import and call the existing bulk upload handler directly, passing our
        # newly-created table_id.  We intercept its JsonResponse to enrich it.
        from core.views.idcard_api import api_idcard_bulk_upload as _bulk_upload

        # Temporarily patch the URL kwargs so the bulk upload sees our table
        bulk_response = _bulk_upload(request, table.id)

        # Parse the JSON body from the upload response
        import json as _json
        try:
            body = _json.loads(bulk_response.content)
        except Exception:
            body = {}

        if bulk_response.status_code == 200 and body.get('success'):
            body['table_id'] = table.id
            body['table_name'] = table.name
            body['fields_created'] = len(fields)
            body['message'] = (
                f'Table "{table.name}" created with {len(fields)} fields. '
                + (body.get('message') or
                   f'{body.get("cards_created", 0)} cards imported.')
            )
            return JsonResponse(body)
        else:
            # Upload failed — clean up: delete the empty table
            try:
                table.delete()
            except Exception:
                pass
            return bulk_response

    except Exception as exc:
        logger.exception("Bulk upload after table creation failed: %s", exc)
        # Clean up table
        try:
            table.delete()
        except Exception:
            pass
        return JsonResponse({
            'success': False,
            'message': 'Table was created but data import failed. Please try again.'
        }, status=500)


# ==================== ID CARD API ENDPOINTS ====================

@require_http_methods(["GET"])
@api_require_any_authenticated
def api_idcard_list(request, table_id):
    """API endpoint to list ID Cards for a table with pagination support for lazy loading.
    
    Supports server-side filtering via query params:
        search  - full-text search on field_data
        class   - exact class filter on field_data
        section - exact section filter on field_data
        sort    - sort order: sr-asc, sr-desc, name-asc, name-desc, date-new, date-old
        image_column    - image field name for image sort filter
        image_condition - complete, pending, or incomplete
    """
    from django.db.models.fields.json import KeyTextTransform
    from django.db.models import Q

    table, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    status_filter = request.GET.get('status', None)
    
    # Check status-specific list permission (via single authority)
    if status_filter:
        required_perm = PermissionService.STATUS_LIST_PERM_MAP.get(status_filter)
        if required_perm and not PermissionService.has(request.user, required_perm):
            return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
    
    try:
        offset = max(0, int(request.GET.get('offset', 0)))
        limit = min(500, max(1, int(request.GET.get('limit', 100))))
    except (ValueError, TypeError):
        offset, limit = 0, 100

    # Server-side search & filters
    search = request.GET.get('search', '').strip()
    class_filter = request.GET.get('class', '').strip()
    section_filter = request.GET.get('section', '').strip()
    sort_order = request.GET.get('sort', 'sr-asc').strip()
    image_column = request.GET.get('image_column', '').strip()
    image_condition = request.GET.get('image_condition', '').strip()
    from_date = request.GET.get('from', '').strip()
    to_date = request.GET.get('to', '').strip()

    result = IDCardService.list_cards(
        table_id, status_filter, offset, limit,
        search=search, class_filter=class_filter, section_filter=section_filter,
        sort_order=sort_order, image_column=image_column, image_condition=image_condition,
        from_date=from_date, to_date=to_date,
    )
    return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)


@require_http_methods(["GET"])
@api_require_any_authenticated
def api_idcard_cards_json(request, table_id):
    """JSON endpoint for virtual table rendering.

    Returns lightweight card data with full filter/pagination support,
    matching the same query logic as the idcard_actions page view.

    Query params:
        status   – filter by status (pending, verified, approved, download, pool, reprint)
        offset   – pagination offset (default 0)
        limit    – page size, 1-500 (default 100)
        search   – full-text search on field_data
        class    – class filter on field_data
        section  – section filter on field_data
        from     – datetime lower bound (download status only)
        to       – datetime upper bound (download status only)

    Response shape:
        {
            "success": true,
            "total": 1234,
            "offset": 0,
            "limit": 100,
            "has_more": true,
            "results": [
                {
                    "id": 5,
                    "sr_no": 1,
                    "status": "pending",
                    "status_display": "Pending",
                    "field_data": {"Name": "...", "PHOTO": "..."},
                    "ordered_fields": [
                        {"name": "Name", "type": "text", "value": "John"},
                        {"name": "PHOTO", "type": "image", "value": "adarshimg/...jpg",
                         "thumb": "adarshimg/thumbs/...jpg"}
                    ],
                    "updated_at": "20-Feb-2026 14:30",
                    "updated_at_iso": "2026-02-20T14:30:00+05:30",
                    "downloaded_at": null,
                    "deleted_at": null
                }
            ]
        }
    """
    from django.utils.timezone import localtime, make_aware, is_naive
    from datetime import datetime as dt

    table, err = _check_client_scope_by_table(request.user, table_id)
    if err:
        return err

    status_filter = request.GET.get('status', None)

    # Check status-specific list permission
    if status_filter:
        required_perm = PermissionService.STATUS_LIST_PERM_MAP.get(status_filter)
        if required_perm and not PermissionService.has(request.user, required_perm):
            return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)

    # Pagination — supports cursor-based (preferred) and offset (legacy)
    cursor = request.GET.get('cursor', '').strip()
    try:
        offset = max(0, int(request.GET.get('offset', 0)))
        limit = min(500, max(1, int(request.GET.get('limit', 100))))
    except (ValueError, TypeError):
        offset, limit = 0, 100

    # Base queryset — show recently moved cards first, defer heavy photo ImageField column
    if status_filter == 'download':
        qs = IDCard.objects.filter(table=table).defer('photo').order_by('-downloaded_at', '-id')
    elif status_filter == 'pool':
        qs = IDCard.objects.filter(table=table).defer('photo').order_by('-deleted_at', '-id')
    else:
        qs = IDCard.objects.filter(table=table).defer('photo').order_by('-updated_at', '-id')

    if status_filter and status_filter in IDCardService.VALID_STATUSES:
        qs = qs.filter(status=status_filter)

    # Search & filter on field_data JSON
    search = request.GET.get('search', '').strip()
    class_filter = request.GET.get('class', '').strip()
    section_filter = request.GET.get('section', '').strip()
    if search:
        qs = qs.filter(field_data__icontains=search)

    # Class/section filter with canonical normalization
    if class_filter or section_filter:
        from django.db.models.fields.json import KeyTextTransform
        class_field_name, section_field_name = _get_class_section_field_names(table)
        if class_filter and class_field_name:
            qs = _build_class_filter_q(qs, class_filter, class_field_name)
        if section_filter and section_field_name:
            qs = qs.annotate(_sec=KeyTextTransform(section_field_name, 'field_data'))
            qs = qs.filter(_sec__iexact=section_filter)

    # Server-side image filter (column + condition)
    image_column = request.GET.get('image_column', '').strip()
    image_condition = request.GET.get('image_condition', '').strip()
    if image_column and image_condition in ('complete', 'pending', 'incomplete'):
        from django.db.models.fields.json import KeyTextTransform
        from django.db.models import Q
        qs = qs.annotate(_img=KeyTextTransform(image_column, 'field_data'))
        if image_condition == 'complete':
            qs = qs.exclude(_img__isnull=True).exclude(_img='').exclude(_img='NOT_FOUND')
            qs = qs.exclude(_img__startswith='PENDING:')
        elif image_condition == 'pending':
            qs = qs.filter(_img__startswith='PENDING:')
        elif image_condition == 'incomplete':
            qs = qs.filter(Q(_img__isnull=True) | Q(_img='') | Q(_img='NOT_FOUND'))

    # DateTime range (download list)
    if status_filter == 'download':
        from_date = request.GET.get('from', '').strip()
        to_date = request.GET.get('to', '').strip()
        if from_date:
            try:
                from_dt = dt.fromisoformat(from_date)
                from_dt = make_aware(from_dt) if is_naive(from_dt) else from_dt
                qs = qs.filter(downloaded_at__gte=from_dt)
            except (ValueError, TypeError):
                pass
        if to_date:
            try:
                to_dt = dt.fromisoformat(to_date)
                to_dt = make_aware(to_dt) if is_naive(to_dt) else to_dt
                qs = qs.filter(downloaded_at__lte=to_dt)
            except (ValueError, TypeError):
                pass

    total = qs.count()

    # Cursor-based pagination: WHERE id < cursor ORDER BY id DESC LIMIT N
    # Falls back to offset pagination if cursor not provided (backward compat)
    if cursor:
        try:
            cursor_id = int(cursor)
            cards = list(qs.filter(id__lt=cursor_id)[:limit + 1])
        except (ValueError, TypeError):
            cards = list(qs[offset:offset + limit + 1])
    else:
        cards = list(qs[offset:offset + limit + 1])

    has_more = len(cards) > limit
    if has_more:
        cards = cards[:limit]
    next_cursor = cards[-1].id if cards and has_more else None

    # Build ordered fields with reordered display order
    reordered_fields = BaseService.reorder_fields_for_display(table.fields or [])

    def _thumb(path):
        """Replicate get_thumbnail_path template filter."""
        if not path or path == 'NOT_FOUND' or path.startswith('PENDING:'):
            return path
        # Reject values that don't look like real file paths (no extension)
        if '.' not in path:
            return ''
        try:
            parts = path.replace('\\', '/').split('/')
            if len(parts) >= 2:
                return parts[0] + '/thumbs/' + '/'.join(parts[1:])
            return 'thumbs/' + path
        except Exception:
            return path

    results = []
    # sr_no base: for cursor mode, use offset param if provided, otherwise 0
    sr_base = offset if not cursor else offset
    for idx, card in enumerate(cards):
        fd = card.field_data or {}
        fd_upper = {k.upper(): v for k, v in fd.items()}

        ordered = []
        for field in reordered_fields:
            fname = field['name']
            ftype = field.get('type', 'text')
            is_img = BaseService.is_image_field(field)
            if is_img:
                ftype = 'image'
            val = fd.get(fname, '') or fd_upper.get(fname.upper(), '')
            # Legacy photo fallback
            if not val and fname.upper() == 'PHOTO' and card.photo:
                try:
                    val = card.photo.name or card.photo.url
                except Exception:
                    pass
            entry = {'name': fname, 'type': ftype, 'value': val}
            if is_img:
                entry['thumb'] = _thumb(val) if val else ''
            ordered.append(entry)

        results.append({
            'id': card.id,
            'sr_no': sr_base + idx + 1,
            'status': card.status,
            'status_display': card.get_status_display(),
            'field_data': fd,
            'ordered_fields': ordered,
            'updated_at': localtime(card.updated_at).strftime('%d-%b-%Y %H:%M') if card.updated_at else None,
            'updated_at_iso': card.updated_at.isoformat() if card.updated_at else None,
            'downloaded_at': localtime(card.downloaded_at).strftime('%d-%b-%Y %H:%M') if card.downloaded_at else None,
            'deleted_at': localtime(card.deleted_at).strftime('%d-%b-%Y %H:%M') if card.deleted_at else None,
        })

    return JsonResponse({
        'success': True,
        'total': total,
        'offset': offset,
        'limit': limit,
        'has_more': has_more,
        'next_cursor': next_cursor,
        'results': results,
    })


@require_http_methods(["GET"])
@api_require_any_authenticated
def api_idcard_all_ids(request, table_id):
    """API endpoint to get all card IDs for a table (for Select All functionality).
    Supports search, class, and section filter params so Select All respects active filters."""
    table, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    status_filter = request.GET.get('status', None)
    
    # Check status-specific list permission (via single authority)
    if status_filter:
        required_perm = PermissionService.STATUS_LIST_PERM_MAP.get(status_filter)
        if required_perm and not PermissionService.has(request.user, required_perm):
            return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
    
    # Pass through the same filters the main view uses
    search = request.GET.get('search', '').strip()
    class_filter = request.GET.get('class', '').strip()
    section_filter = request.GET.get('section', '').strip()
    from_date = request.GET.get('from', '').strip()
    to_date = request.GET.get('to', '').strip()
    image_column = request.GET.get('image_column', '').strip()
    image_condition = request.GET.get('image_condition', '').strip()
    
    result = IDCardService.get_all_card_ids(
        table_id, status_filter,
        search=search, class_filter=class_filter, section_filter=section_filter,
        from_date=from_date, to_date=to_date,
        image_column=image_column, image_condition=image_condition,
    )
    return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)


@require_http_methods(["GET"])
@api_require_any_authenticated
def api_idcard_filter_options(request, table_id):
    """Return distinct class/section values for filter dropdowns.

    Groups class variants by their canonical form (normalize_class_value).
    E.g. 'KG-I', 'KGI', 'KG1', 'kgI' all map to canonical 'KG1' and show
    the most-used raw format as the display label.

    Response shape:
        class_values:   [{value: "KG1", display: "KG-I"}, ...]
        section_values: ["A", "B", ...]
    """
    from django.db.models.fields.json import KeyTextTransform
    from django.db.models.functions import Cast
    from django.db.models import CharField, Count
    from core.utils.field_utils import (
        CLASS_ORDER, CLASS_ORDER_UNKNOWN, normalize_class_value,
    )
    from collections import defaultdict

    table, err = _check_client_scope_by_table(request.user, table_id)
    if err:
        return err

    status_filter = request.GET.get('status', '').strip()

    qs = IDCard.objects.filter(table=table)
    if status_filter and status_filter in IDCardService.VALID_STATUSES:
        qs = qs.filter(status=status_filter)

    class_field_name, section_field_name = _get_class_section_field_names(table)

    class_values = []
    section_values = []

    if class_field_name:
        # Get distinct raw values WITH counts
        raw_with_counts = (
            qs.annotate(_cv=Cast(KeyTextTransform(class_field_name, 'field_data'), CharField()))
            .exclude(_cv__isnull=True).exclude(_cv='')
            .order_by()
            .values('_cv')
            .annotate(cnt=Count('id'))
        )

        # Group by canonical form → pick most common raw as display
        groups = defaultdict(list)  # canonical → [(raw, count)]
        for entry in raw_with_counts:
            raw = entry['_cv'].strip()
            canonical = normalize_class_value(raw)
            groups[canonical].append((raw, entry['cnt']))

        for canonical, variants in groups.items():
            best_display = max(variants, key=lambda x: x[1])[0]
            total_count = sum(v[1] for v in variants)
            class_values.append({
                'value': canonical,
                'display': best_display,
                'count': total_count,
            })

        # Sort by class order
        class_values.sort(
            key=lambda x: (CLASS_ORDER.get(x['value'], CLASS_ORDER_UNKNOWN), x['value'])
        )

    if section_field_name:
        section_values = sorted(
            [
                str(v) for v in
                qs.annotate(_sv=Cast(KeyTextTransform(section_field_name, 'field_data'), CharField()))
                .exclude(_sv__isnull=True).exclude(_sv='')
                .order_by()
                .values_list('_sv', flat=True).distinct()
                if v is not None
            ],
        )

    return JsonResponse({
        'success': True,
        'class_values': class_values,
        'section_values': list(section_values),
        'class_field': class_field_name,
        'section_field': section_field_name,
    })


@require_http_methods(["POST"])
@api_require_permission('perm_idcard_add')
def api_idcard_create(request, table_id):
    """API endpoint to create a new ID Card with file upload support.

    View responsibility: parse HTTP request → delegate to IDCardService.
    """
    _tbl, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    try:
        # Parse field_data and image_files from either multipart or JSON
        if request.content_type and 'multipart/form-data' in request.content_type:
            field_data = json.loads(request.POST.get('field_data', '{}'))
            # CRITICAL: dict(request.FILES) returns lists (MultiValueDict internals).
            # Use dict comprehension with [] access to get actual file objects.
            image_files = {key: request.FILES[key] for key in request.FILES}
            legacy_photo_file = request.FILES.get('photo')
        else:
            data = json.loads(request.body)
            field_data = data.get('field_data', {})
            image_files = None
            legacy_photo_file = None

        result = IDCardService.create_card(
            table_id=table_id,
            field_data=field_data,
            image_files=image_files,
            uploaded_by=request.user if request.user.is_authenticated else None,
            legacy_photo_file=legacy_photo_file,
        )

        if result.success:
            return JsonResponse({
                'success': True,
                'message': result.message,
                'card': result.data['card'],
            })
        return JsonResponse({'success': False, 'message': result.message}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data!'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["GET"])
@api_require_permission('perm_idcard_info')
def api_idcard_get(request, card_id):
    """API endpoint to get a single ID Card"""
    card, err = _check_client_scope_by_card(request.user, card_id)
    if err: return err
    result = IDCardService.get_card(card_id)
    if result.success:
        return JsonResponse({'success': True, 'card': result.data['card']})
    return JsonResponse({'success': False, 'message': result.message}, status=400)


@require_http_methods(["POST", "PUT"])
@api_require_permission('perm_idcard_edit')
def api_idcard_update(request, card_id):
    """API endpoint to update an ID Card with file upload support.

    View responsibility: parse HTTP request, scope/readonly gates → delegate
    to IDCardService.update_card (which handles concurrency, images, save).
    """
    _card, err = _check_client_scope_by_card(request.user, card_id)
    if err: return err
    # Client/client_staff cannot edit cards in approved/download/reprint
    if _is_client_readonly(request.user, _card.status):
        return _client_readonly_response()
    try:
        # Parse request into service-friendly args
        if request.content_type and 'multipart/form-data' in request.content_type:
            field_data = json.loads(request.POST.get('field_data', '{}'))
            expected_updated_at = request.POST.get('expected_updated_at', None)
            # CRITICAL: dict(request.FILES) returns lists (MultiValueDict internals).
            # Use dict comprehension with [] access to get actual file objects.
            image_files = {key: request.FILES[key] for key in request.FILES}
            legacy_photo_file = request.FILES.get('photo')
        else:
            data = json.loads(request.body)
            field_data = data.get('field_data')
            expected_updated_at = data.get('expected_updated_at', None)
            image_files = None
            legacy_photo_file = None

        result = IDCardService.update_card(
            card_id=card_id,
            field_data=field_data,
            image_files=image_files,
            uploaded_by=request.user if request.user.is_authenticated else None,
            expected_updated_at=expected_updated_at,
            legacy_photo_file=legacy_photo_file,
        )

        if result.success:
            card_data = result.data['card']
            return JsonResponse({
                'success': True,
                'message': result.message,
                'card': {
                    'id': card_data['id'],
                    'field_data': card_data['field_data'],
                    'photo': card_data.get('photo'),
                    'status': card_data['status'],
                    'status_display': card_data.get('status_display'),
                    'updated_at': card_data.get('updated_at_iso'),
                }
            })

        # Concurrency conflict → 409
        if result.data and result.data.get('conflict'):
            return JsonResponse({
                'success': False,
                'message': result.message,
                'conflict': True,
                'server_updated_at': result.data['server_updated_at'],
            }, status=409)

        return JsonResponse({'success': False, 'message': result.message}, status=400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data!'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["DELETE", "POST"])
@api_require_permission('perm_idcard_delete')
def api_idcard_delete(request, card_id):
    """API endpoint to delete an ID Card"""
    _card, err = _check_client_scope_by_card(request.user, card_id)
    if err: return err
    # Client/client_staff cannot delete cards in approved/download/reprint
    if _is_client_readonly(request.user, _card.status):
        return _client_readonly_response()
    try:
        result = IDCardService.delete_card(card_id)
        return JsonResponse(
            {'success': result.success, 'message': result.message},
            status=200 if result.success else 400
        )
    except Exception as e:
        logger.exception("Card delete error: %s", e)
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["POST"])
@api_require_permission('perm_idcard_edit')
def api_idcard_update_field(request, card_id):
    """API endpoint to update a single field on an ID Card (for inline editing)"""
    _card, err = _check_client_scope_by_card(request.user, card_id)
    if err: return err
    # Client/client_staff cannot edit cards in approved/download/reprint
    if _is_client_readonly(request.user, _card.status):
        return _client_readonly_response()
    try:
        data = json.loads(request.body)
        field = data.get('field')
        value = data.get('value', '')
        
        result = IDCardService.update_single_field(card_id, field, value)
        return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data!'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["POST"])
@api_require_any_authenticated
def api_idcard_change_status(request, card_id):
    """API endpoint to change an ID Card's status.
    
    Delegates entirely to WorkflowService.transition() which enforces:
    transition matrix, permissions, mandatory fields, image gate, client-readonly, activity log.
    """
    card, err = _check_client_scope_by_card(request.user, card_id)
    if err: return err
    try:
        data = json.loads(request.body)
        new_status = data.get('status')

        from ..services.workflow_service import WorkflowService
        result = WorkflowService.transition(card, new_status, user=request.user, request=request)
        return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data!'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["POST"])
@api_require_any_authenticated
def api_idcard_bulk_status(request, table_id):
    """API endpoint to change status of multiple ID Cards.
    
    Delegates entirely to WorkflowService.bulk_transition() which enforces:
    transition matrix, permissions, mandatory fields, image gate, client-readonly, activity log.
    """
    _tbl, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    try:
        data = json.loads(request.body)
        card_ids = data.get('card_ids', [])
        new_status = data.get('status')
        
        if not new_status:
            return JsonResponse({'success': False, 'message': 'Status is required'}, status=400)
        if not card_ids:
            return JsonResponse({'success': False, 'message': 'No cards selected'}, status=400)

        from ..services.workflow_service import WorkflowService
        result = WorkflowService.bulk_transition(
            _tbl, card_ids, new_status, user=request.user, request=request
        )
        return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data!'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["POST"])
@api_require_any_authenticated
def api_idcard_bulk_delete(request, table_id):
    """API endpoint to delete multiple ID Cards.
    When delete_all=True, requires perm_delete_all_idcard + 6-digit confirmation_code.
    When delete_all=False (selected cards), requires perm_idcard_delete_from_pool.
    """
    _tbl, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    try:
        data = json.loads(request.body)
        card_ids = data.get('card_ids', [])
        delete_all = data.get('delete_all', False)
        
        if not delete_all and not card_ids:
            return JsonResponse({'success': False, 'message': 'No cards selected'}, status=400)
        
        # Check appropriate permission
        if delete_all:
            if not PermissionService.has(request.user, 'perm_delete_all_idcard'):
                return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
        else:
            if not PermissionService.has(request.user, 'perm_idcard_delete_from_pool'):
                return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
        
        # Secure confirmation for delete-all
        if delete_all:
            confirmation_code = data.get('confirmation_code', '')
            session_key = f'delete_all_code_{table_id}'
            expected_code = request.session.get(session_key)
            
            if not expected_code:
                return JsonResponse({
                    'success': False,
                    'message': 'No confirmation code generated. Please request a new code.'
                }, status=400)
            
            if str(confirmation_code) != str(expected_code):
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid confirmation code. Delete aborted.'
                }, status=403)
            
            # Code verified — clear it so it can't be reused
            del request.session[session_key]
            request.session.modified = True
        
        result = IDCardService.bulk_delete(table_id, card_ids, delete_all)
        if result.success:
            count = result.data.get('deleted_count', len(card_ids))
            target_label = 'all cards' if delete_all else f'{count} card(s)'
            ActivityService.log_bulk_delete(request, target_label, count)
        return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data!'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["POST"])
@api_require_permission('perm_delete_all_idcard')
def api_generate_delete_code(request, table_id):
    """Generate a 6-digit confirmation code for delete-all, stored in session."""
    import secrets
    _tbl, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    try:
        table = _tbl  # Reuse already-fetched table from scope check
        total = IDCard.objects.filter(table=table).count()
        
        code = str(secrets.randbelow(900000) + 100000)
        request.session[f'delete_all_code_{table_id}'] = code
        request.session.modified = True
        
        return JsonResponse({
            'success': True,
            'code': code,
            'table_name': table.name,
            'total_cards': total,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["POST"])
@api_require_permission('perm_idcard_upgrade_all')
def api_generate_upgrade_code(request, table_id):
    """Generate a 6-digit confirmation code for upgrade-all-classes, stored in session."""
    import secrets
    _tbl, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    try:
        table = _tbl  # Reuse already-fetched table from scope check
        download_count = IDCard.objects.filter(table=table, status='download').count()

        code = str(secrets.randbelow(900000) + 100000)
        request.session[f'upgrade_all_code_{table_id}'] = code
        request.session.modified = True

        return JsonResponse({
            'success': True,
            'code': code,
            'table_name': table.name,
            'download_count': download_count,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["POST"])
@api_require_permission('perm_idcard_upgrade_all')
def api_upgrade_all_classes(request, table_id):
    """
    Upgrade the class field value for all cards in the 'download' list.
    Each class value is bumped to the next level (e.g. V → VI).
    Cards already at XII remain unchanged.
    Only affects cards with status='download'.
    Requires 6-digit confirmation code.
    """
    _tbl, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    try:
        # Verify confirmation code (session validation stays in view — request-scoped)
        data = json.loads(request.body) if request.body else {}
        confirmation_code = data.get('confirmation_code', '')
        expected_code = request.session.get(f'upgrade_all_code_{table_id}', '')

        if not expected_code or confirmation_code != expected_code:
            return JsonResponse({
                'success': False,
                'message': 'Invalid or expired confirmation code. Please try again.'
            }, status=400)

        # Clear the code after use
        request.session.pop(f'upgrade_all_code_{table_id}', None)
        request.session.modified = True

        # Delegate to service layer
        result = IDCardService.upgrade_all_classes(table_id)
        if not result.success:
            return JsonResponse({'success': False, 'message': result.message}, status=400)

        if result.data.get('upgraded', 0) > 0:
            ActivityService.log_bulk_upgrade(
                request, result.data['upgraded'], result.data.get('client_name', '')
            )

        return JsonResponse({
            'success': True,
            'message': result.message,
            'upgraded': result.data['upgraded'],
            'skipped': result.data['skipped'],
            'total': result.data['total'],
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["GET"])
@api_require_any_authenticated
def api_idcard_search(request, table_id):
    """API endpoint to search ID Cards across all statuses"""
    _tbl, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    query = request.GET.get('q', '').strip()
    result = IDCardService.search_cards(table_id, query)
    return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)


@require_http_methods(["GET"])
@api_require_any_authenticated
def api_table_status_counts(request, table_id):
    """API endpoint to get status counts for a table"""
    _tbl, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    try:
        table = _tbl  # Reuse already-fetched table from scope check
        status_counts = IDCardService.get_status_counts(table)
        
        return JsonResponse({
            'success': True,
            'status_counts': status_counts
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["POST"])
@api_require_permission('perm_idcard_bulk_upload')
def api_idcard_bulk_upload(request, table_id):
    """API endpoint to bulk upload ID Cards from XLSX/CSV file with fuzzy matching and optional ZIP photo upload.
    
    Uses disk-based image storage for large ZIPs to prevent OOM.
    Small datasets (<50MB total) are kept in RAM for speed.
    Row processing is unified across XLSX and CSV via BulkUploadService.
    """
    _tbl, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    # Double-click guard: prevent duplicate uploads from rapid form submissions
    lock_key = f'bulk_upload_lock:{request.user.id}:{table_id}'
    if not django_cache.add(lock_key, 1, 300):
        return JsonResponse({'success': False, 'message': 'Upload already in progress. Please wait.'}, status=429)
    
    # Import the disk-backed image store and helpers
    from ..services.bulk_upload_service import (
        DiskBackedImageStore, extract_zip_to_store, process_data_rows,
        MAX_BULK_ROWS, BULK_BATCH_SIZE,
    )
    
    # Track all image stores for cleanup on exit
    _all_stores = []
    
    try:
        import openpyxl
        from io import BytesIO
        import re
        import zipfile
        import os
        import shutil
        from django.core.files.storage import default_storage
        from django.core.files.base import ContentFile

        # Pre-flight disk space check: require at least 500 MB free
        try:
            disk = shutil.disk_usage(settings.MEDIA_ROOT)
            if disk.free < 500 * 1024 * 1024:  # 500 MB
                return JsonResponse({
                    'success': False,
                    'message': 'Insufficient disk space. Please contact your administrator.'
                }, status=507)
        except Exception:
            pass  # Non-critical — proceed if check fails
        
        table = get_object_or_404(IDCardTable.objects.select_related('group__client'), id=table_id)
        
        if 'file' not in request.FILES:
            return JsonResponse({'success': False, 'message': 'No file uploaded!'}, status=400)
        
        uploaded_file = request.FILES['file']
        file_name = uploaded_file.name.lower()
        file_size = uploaded_file.size
        
        # Get image field names from table using BaseService
        image_field_names = BaseService.get_image_field_names(table.fields)
        
        # ── Extract ZIP images into disk-backed stores ──
        # Each image field gets its own DiskBackedImageStore
        zip_photos_by_field = {}  # { field_name: DiskBackedImageStore }
        
        # Check for multiple ZIP files - one per image field
        zip_field_names_str = request.POST.get('zip_field_names', '[]')
        try:
            zip_field_names = json.loads(zip_field_names_str)
        except (json.JSONDecodeError, TypeError):
            zip_field_names = []
        
        logger.debug("zip_field_names = %s", zip_field_names)
        
        # Process per-field ZIP files
        for field_name in zip_field_names:
            zip_key = f'photos_zip_{field_name}'
            if zip_key in request.FILES:
                photos_zip_file = request.FILES[zip_key]
                store = DiskBackedImageStore()
                _all_stores.append(store)
                count = extract_zip_to_store(photos_zip_file, store)
                if count > 0:
                    zip_photos_by_field[field_name] = store
                    logger.debug("Field '%s' extracted %d images", field_name, count)
        
        # Legacy: single photos_zip (backward compatibility)
        if not zip_photos_by_field and 'photos_zip' in request.FILES:
            photos_zip_file = request.FILES['photos_zip']
            first_image_field = image_field_names[0] if image_field_names else 'PHOTO'
            store = DiskBackedImageStore()
            _all_stores.append(store)
            count = extract_zip_to_store(photos_zip_file, store)
            if count > 0:
                zip_photos_by_field[first_image_field] = store
        
        # Unified ZIP files (images auto-matched to all columns)
        unified_zip_photos = DiskBackedImageStore()
        _all_stores.append(unified_zip_photos)
        
        try:
            unified_zip_count = min(int(request.POST.get('unified_zip_count', 0)), 20)
        except (ValueError, TypeError):
            unified_zip_count = 0
        
        for i in range(unified_zip_count):
            zip_key = f'unified_zip_{i}'
            if zip_key in request.FILES:
                extract_zip_to_store(request.FILES[zip_key], unified_zip_photos)
        
        logger.debug("unified_zip_photos count = %d", len(unified_zip_photos))
        
        # Get client for ImageService operations
        client = table.group.client
        
        # Get all table fields
        all_table_fields = table.fields
        table_fields = [f['name'] for f in all_table_fields if not BaseService.is_image_field(f)]
        image_fields = [f['name'] for f in all_table_fields if BaseService.is_image_field(f)]
        
        matched_field_names = []
        
        # Check if frontend sent a manual field mapping
        frontend_mapping_str = request.POST.get('field_mapping', '')
        frontend_mapping = {}
        if frontend_mapping_str:
            try:
                frontend_mapping = json.loads(frontend_mapping_str)
                if not isinstance(frontend_mapping, dict):
                    frontend_mapping = {}
            except (json.JSONDecodeError, TypeError):
                frontend_mapping = {}
        
        # ── Parse file (XLSX/XLS/CSV) ──
        if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
            rows_data, headers, header_to_field, image_ref_columns, matched_field_names, parse_error = \
                _parse_excel_file(uploaded_file, file_name, table_fields, image_fields,
                                  frontend_mapping, all_table_fields)
            if parse_error:
                return parse_error
            is_csv = False
        elif file_name.endswith('.csv'):
            rows_data, headers, header_to_field, image_ref_columns, matched_field_names, parse_error = \
                _parse_csv_file(uploaded_file, table_fields, image_fields,
                                frontend_mapping, all_table_fields)
            if parse_error:
                return parse_error
            is_csv = True
        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid file format! Please upload .xlsx, .xls, or .csv file.'
            }, status=400)
        
        if not header_to_field:
            return JsonResponse({
                'success': False,
                'message': f'No matching columns found! Expected columns: {", ".join(table_fields)}'
            }, status=400)
        
        if len(rows_data) > MAX_BULK_ROWS:
            return JsonResponse({
                'success': False,
                'message': f'File has {len(rows_data)} rows. Maximum allowed is {MAX_BULK_ROWS}.'
            }, status=400)
        
        # Reverse rows so first Excel row gets highest DB id (preserves order in -id display)
        rows_data = list(reversed(rows_data))
        
        # ── Process rows using unified service ──
        result = process_data_rows(
            rows=rows_data,
            header_to_field=header_to_field,
            image_ref_columns=image_ref_columns,
            image_fields=image_fields,
            all_table_fields=all_table_fields,
            table=table,
            client=client,
            zip_photos_by_field=zip_photos_by_field,
            unified_zip_photos=unified_zip_photos,
            request_user=request.user,
            is_csv=is_csv,
        )
        
        cards_created = result['cards_created']
        total_photos_matched = result['total_photos_matched']
        errors = result['errors']
        
        # Return result
        photo_msg = f" with {total_photos_matched} photos matched" if total_photos_matched > 0 else ""
        response = {
            'success': True,
            'message': f'Successfully created {cards_created} ID cards{photo_msg}!',
            'cards_created': cards_created,
            'photos_matched': total_photos_matched,
            'matched_fields': matched_field_names,
        }
        
        if errors:
            response['errors'] = errors[:10]
            response['error_count'] = len(errors)
        
        return JsonResponse(response)
        
    except ImportError:
        return JsonResponse({
            'success': False,
            'message': 'openpyxl library not installed. Run: pip install openpyxl'
        }, status=500)
    except Exception as e:
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)
    finally:
        # Always cleanup disk-backed image stores
        for store in _all_stores:
            try:
                store.cleanup()
            except Exception:
                pass
        django_cache.delete(lock_key)


def _parse_excel_file(uploaded_file, file_name, table_fields, image_fields,
                       frontend_mapping, all_table_fields):
    """Parse Excel file and return (rows_data, headers, header_to_field, image_ref_columns, matched_fields, error_response).
    Returns error_response=None on success, or a JsonResponse on failure."""
    import openpyxl
    from io import BytesIO
    
    try:
        file_content = uploaded_file.read()
        if len(file_content) < 4:
            return None, None, None, None, [], JsonResponse({
                'success': False, 'message': 'File is too small or empty.'
            }, status=400)
        
        magic_bytes = file_content[:4]
        is_zip = magic_bytes[:2] == b'PK'
        is_old_xls = (magic_bytes[0] == 0xD0 and magic_bytes[1] == 0xCF)
        
        headers = []
        rows_data = []
        
        if is_zip or file_name.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(BytesIO(file_content))
                ws = wb.active
                for cell in ws[1]:
                    if cell.value:
                        headers.append(str(cell.value).strip()
                                       .replace('_x000D_', '').replace('_X000D_', '')
                                       .replace('_x000d_', '').replace('\r', ''))
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows_data.append(row)
            except Exception as xlsx_error:
                if not is_zip:
                    is_old_xls = True
                else:
                    raise xlsx_error
        
        if is_old_xls or (file_name.endswith('.xls') and not file_name.endswith('.xlsx') and not headers):
            try:
                import xlrd
                wb = xlrd.open_workbook(file_contents=file_content)
                ws = wb.sheet_by_index(0)
                headers = []
                for col_idx in range(ws.ncols):
                    cell_value = ws.cell_value(0, col_idx)
                    if cell_value:
                        headers.append(str(cell_value).strip())
                rows_data = []
                for row_idx in range(1, ws.nrows):
                    row = []
                    for col_idx in range(ws.ncols):
                        row.append(ws.cell_value(row_idx, col_idx))
                    rows_data.append(tuple(row))
            except ImportError:
                return None, None, None, None, [], JsonResponse({
                    'success': False, 'message': 'xlrd library not installed for .xls files.'
                }, status=400)
            except Exception:
                return None, None, None, None, [], JsonResponse({
                    'success': False, 'message': 'Error reading .xls file. Please check the file format.'
                }, status=400)
        
        if not headers:
            return None, None, None, None, [], JsonResponse({
                'success': False, 'message': 'Could not read headers from Excel file.'
            }, status=400)
    except Exception:
        return None, None, None, None, [], JsonResponse({
            'success': False, 'message': 'Error reading Excel file. Please check the file format.'
        }, status=400)
    
    header_to_field, image_ref_columns, matched_field_names = _map_headers_to_fields(
        headers, table_fields, image_fields, frontend_mapping, all_table_fields, is_csv=False
    )
    
    return rows_data, headers, header_to_field, image_ref_columns, matched_field_names, None


def _parse_csv_file(uploaded_file, table_fields, image_fields,
                     frontend_mapping, all_table_fields):
    """Parse CSV file and return (rows_data, headers, header_to_field, image_ref_columns, matched_fields, error_response)."""
    import csv
    from io import StringIO
    
    try:
        content = uploaded_file.read().decode('utf-8-sig')
        reader = csv.DictReader(StringIO(content))
        csv_headers = reader.fieldnames or []
        rows_data = list(reader)
    except Exception:
        return None, None, None, None, [], JsonResponse({
            'success': False, 'message': 'Error reading CSV file. Please check the file format.'
        }, status=400)
    
    if not csv_headers:
        return None, None, None, None, [], JsonResponse({
            'success': False, 'message': 'Could not read headers from CSV file.'
        }, status=400)
    
    header_to_field, image_ref_columns, matched_field_names = _map_headers_to_fields(
        csv_headers, table_fields, image_fields, frontend_mapping, all_table_fields, is_csv=True
    )
    
    return rows_data, csv_headers, header_to_field, image_ref_columns, matched_field_names, None


def _map_headers_to_fields(headers, table_fields, image_fields, frontend_mapping,
                            all_table_fields, *, is_csv=False):
    """Map file headers to table fields using fuzzy matching or frontend mapping.
    Returns (header_to_field, image_ref_columns, matched_field_names)."""
    header_to_field = {}
    available_fields = table_fields.copy()
    image_ref_columns = {}
    unmatched_image_fields = list(image_fields)
    matched_field_names = []
    
    if frontend_mapping:
        if is_csv:
            for table_field_name, excel_header in frontend_mapping.items():
                if table_field_name in available_fields and excel_header in headers:
                    header_to_field[excel_header] = table_field_name
                    available_fields.remove(table_field_name)
                    matched_field_names.append(table_field_name)
        else:
            header_index = {h: i for i, h in enumerate(headers)}
            for table_field_name, excel_header in frontend_mapping.items():
                if table_field_name in available_fields and excel_header in header_index:
                    idx = header_index[excel_header]
                    header_to_field[idx] = table_field_name
                    available_fields.remove(table_field_name)
                    matched_field_names.append(table_field_name)
        
        # Auto-match image columns
        for idx_or_header in (range(len(headers)) if not is_csv else headers):
            header = headers[idx_or_header] if not is_csv else idx_or_header
            if not header:
                continue
            if not is_csv and idx_or_header in header_to_field:
                continue
            if is_csv and header in header_to_field:
                continue
            matched_img = BaseService.find_best_image_field_match(header, unmatched_image_fields)
            if matched_img:
                image_ref_columns[matched_img] = idx_or_header if not is_csv else header
                unmatched_image_fields.remove(matched_img)
    else:
        # Auto fuzzy matching
        for idx_or_header in (range(len(headers)) if not is_csv else headers):
            header = headers[idx_or_header] if not is_csv else idx_or_header
            if not header:
                continue
            
            matched_img = BaseService.find_best_image_field_match(header, unmatched_image_fields)
            if matched_img:
                image_ref_columns[matched_img] = idx_or_header if not is_csv else header
                unmatched_image_fields.remove(matched_img)
                continue
            
            header_str = header if is_csv else header
            match = BaseService.find_best_field_match(header_str.strip() if is_csv else header, available_fields)
            if match:
                if is_csv:
                    header_to_field[header] = match
                else:
                    header_to_field[idx_or_header] = match
                available_fields.remove(match)
                matched_field_names.append(match)
    
    return header_to_field, image_ref_columns, matched_field_names


@require_http_methods(["POST"])
@api_require_permission('perm_idcard_bulk_reupload')
def api_idcard_reupload_images(request, table_id):
    """
    API endpoint to reupload images from a ZIP file.
    Matches ZIP filenames to card image references (PENDING: or existing paths) and updates them.
    
    Supports:
    - PENDING:reference matching (for cards created without images)
    - Existing image path updates (applies edit naming: original_14 + _HHMMSS)
    - Multiple image fields per card
    - Thumbnail generation for all saved images
    """
    _tbl, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    # Client/client_staff cannot reupload images for tables with approved/download/reprint cards
    if request.user.role in ('client', 'client_staff'):
        has_locked = IDCard.objects.filter(
            table_id=table_id, status__in=_CLIENT_READONLY_STATUSES
        ).exists()
        if has_locked:
            return JsonResponse({
                'success': False,
                'message': 'This table contains cards in approved/download status. Client users cannot reupload images.'
            }, status=403)
    # Double-click guard: prevent duplicate reupload from rapid form submissions
    lock_key = f'reupload_lock:{request.user.id}:{table_id}'
    if not django_cache.add(lock_key, 1, 300):
        return JsonResponse({'success': False, 'message': 'Reupload already in progress. Please wait.'}, status=429)
    try:
        import zipfile
        from django.db import transaction
        
        table = get_object_or_404(IDCardTable.objects.select_related('group__client'), id=table_id)
        client = table.group.client
        
        if 'photos_zip' not in request.FILES:
            return JsonResponse({'success': False, 'message': 'No ZIP file uploaded!'}, status=400)
        
        # Get image field names from table
        image_field_names = BaseService.get_image_field_names(table.fields)
        if not image_field_names:
            return JsonResponse({'success': False, 'message': 'No image fields defined in table!'}, status=400)
        
        # Get target field from request (optional - defaults to first image field)
        target_field = request.POST.get('target_field', image_field_names[0])
        if target_field not in image_field_names:
            target_field = image_field_names[0]
        
        # Extract photos from ZIP — use temp file if available (avoids OOM on large uploads)
        zip_photos = {}  # { normalized_key: { bytes, ext, original_name } }
        
        try:
            zip_file = request.FILES['photos_zip']

            # ZIP size guard
            if hasattr(zip_file, 'size') and zip_file.size > 600 * 1024 * 1024:
                return JsonResponse({'success': False, 'message': 'ZIP file exceeds 600 MB limit.'}, status=400)
            
            # ZIP bomb / nested archive check
            zok, zerr = validate_zip_safety(zip_file)
            if not zok:
                return JsonResponse({'success': False, 'message': zerr}, status=400)

            # Open ZIP directly from file handle (Django spills >10MB to /tmp)
            if hasattr(zip_file, 'temporary_file_path'):
                zf = zipfile.ZipFile(zip_file.temporary_file_path(), 'r')
            else:
                zip_file.seek(0)
                zf = zipfile.ZipFile(zip_file, 'r')

            with zf:
                for zip_info in zf.infolist():
                    if zip_info.is_dir():
                        continue
                    
                    file_in_zip = zip_info.filename
                    base_name = os.path.basename(file_in_zip)
                    name_without_ext = os.path.splitext(base_name)[0]
                    ext = os.path.splitext(base_name)[1].lower()
                    
                    if ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']:
                        try:
                            image_bytes = zf.read(zip_info.filename)
                            is_valid, error_msg = validate_image_bytes(image_bytes)
                            if is_valid:
                                normalized_key = BaseService.normalize_image_identifier(name_without_ext)
                                if normalized_key:
                                    # Deterministic: if duplicate key, keep alphabetically-first filename
                                    existing = zip_photos.get(normalized_key)
                                    if existing is None or base_name < existing['original_name']:
                                        zip_photos[normalized_key] = {
                                            'bytes': image_bytes,
                                            'ext': ext,
                                            'original_name': base_name
                                        }
                        except Exception:
                            continue
        except Exception as zip_error:
            logger.exception('Error reading ZIP file: %s', zip_error)
            return JsonResponse({'success': False, 'message': 'Error reading ZIP file. Please check the file and try again.'}, status=400)
        
        if not zip_photos:
            return JsonResponse({'success': False, 'message': 'No valid images found in ZIP file!'}, status=400)
        
        logger.debug("Reupload: %d images extracted from ZIP, keys: %s", len(zip_photos), list(zip_photos.keys())[:10])
        
        # Get cards — scoped to selected IDs if provided, else all in table for current status
        card_ids = []
        if 'card_ids' in request.POST:
            try:
                card_ids = json.loads(request.POST.get('card_ids', '[]'))
            except (json.JSONDecodeError, TypeError):
                card_ids = []
        
        # Filter out empty/falsy values
        card_ids = [int(cid) for cid in card_ids if cid and str(cid).strip().isdigit()] if card_ids else []
        
        if card_ids:
            cards_qs = IDCard.objects.filter(table=table, id__in=card_ids).order_by('id')
        else:
            # No specific IDs — reupload to ALL cards in this table (filtered by status if provided)
            status_filter = request.POST.get('status', '')
            if status_filter and status_filter in BaseService.VALID_STATUSES:
                cards_qs = IDCard.objects.filter(table=table, status=status_filter).order_by('id')
            else:
                cards_qs = IDCard.objects.filter(table=table).order_by('id')
        
        updated_count = 0
        matched_count = 0
        errors = []
        
        # Process cards in batches to avoid long-running transactions
        # (SQLite locks the entire DB during a transaction; smaller batches
        # reduce lock duration and prevent "database is locked" errors)
        REUPLOAD_BATCH_SIZE = 50
        all_card_ids = list(cards_qs.values_list('id', flat=True))
        batch_counter = 0
        
        for batch_start in range(0, len(all_card_ids), REUPLOAD_BATCH_SIZE):
            batch_ids = all_card_ids[batch_start:batch_start + REUPLOAD_BATCH_SIZE]
            
            with transaction.atomic():
                batch_cards = IDCard.objects.filter(id__in=batch_ids).order_by('id')
                
                for card in batch_cards:
                    field_data = card.field_data or {}
                    card_updated = False
                    
                    for img_field in image_field_names:
                        current_value = field_data.get(img_field, '')
                        
                        # Determine what to match against
                        match_key = None
                        existing_path = None
                        
                        if current_value.startswith('PENDING:'):
                            # Extract the reference from PENDING:reference
                            match_key = BaseService.normalize_image_identifier(current_value[8:])
                        elif current_value and current_value not in ('NOT_FOUND', ''):
                            # Has existing image - extract filename for matching
                            existing_path = current_value
                            existing_filename = os.path.splitext(os.path.basename(current_value))[0]
                            match_key = BaseService.normalize_image_identifier(existing_filename)
                        else:
                            # No current value - skip unless we want to match by card data
                            # Could extend to match by NAME or other field values
                            continue
                        
                        if not match_key:
                            continue
                        
                        # Try to find matching photo in ZIP
                        if match_key in zip_photos:
                            photo_info = zip_photos[match_key]
                            matched_count += 1
                            
                            try:
                                batch_counter += 1
                                
                                # Use single-authority entry point
                                if existing_path:
                                    result = ImageService.replace_image(
                                        image_bytes=photo_info['bytes'],
                                        client=client,
                                        field_name=img_field,
                                        existing_path=existing_path,
                                        card=card,
                                        batch_counter=batch_counter,
                                        original_ext=photo_info['ext'],
                                    )
                                else:
                                    result = ImageService.save_new_image(
                                        image_bytes=photo_info['bytes'],
                                        client=client,
                                        field_name=img_field,
                                        card=card,
                                        batch_counter=batch_counter,
                                        original_ext=photo_info['ext'],
                                    )
                                
                                if result.success and result.data.get('final_value'):
                                    field_data[img_field] = result.data['final_value']
                                    card_updated = True
                                    logger.debug("Reupload: Card %s field %s updated to %s", 
                                               card.pk, img_field, result.data['final_value'])
                                else:
                                    errors.append(f"Card {card.pk}: Failed to save {img_field} - {result.message}")
                            except Exception as save_err:
                                errors.append(f"Card {card.pk}: Error saving {img_field} - {str(save_err)}")
                    
                    if card_updated:
                        card.field_data = field_data
                        card.save()
                        updated_count += 1
        
        # Build response
        result_msg = f"Updated {updated_count} cards with {matched_count} images matched"
        response = {
            'success': True,
            'message': result_msg,
            'updated_count': updated_count,
            'matched_count': matched_count,
            'zip_images_count': len(zip_photos),
        }
        
        if errors:
            response['errors'] = errors[:10]
            response['error_count'] = len(errors)
        
        return JsonResponse(response)
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)
    finally:
        django_cache.delete(lock_key)


# ==================== MODALS HTML (Lazy Load) ====================

@require_http_methods(["GET"])
@api_require_any_authenticated
def api_idcard_modals_html(request, table_id):
    """Return rendered modals.html partial for lazy-loading.
    Used by modal-loader.js to inject modals on first user interaction
    instead of pre-rendering them on every page load.
    """
    from django.template.loader import render_to_string
    from django.http import HttpResponse

    table, err = _check_client_scope_by_table(request.user, table_id)
    if err:
        return err

    html = render_to_string('partials/idcard/modals.html', {'table': table}, request=request)
    return HttpResponse(html, content_type='text/html; charset=utf-8')